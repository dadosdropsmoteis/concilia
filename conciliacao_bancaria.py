"""
Dashboard de Conciliação Bancária
OFX (Extrato Bancário) x XLS (Intermediadora / Rede)
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
import re

# ─────────────────────────────────────────────
# CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Conciliação Bancária",
    page_icon="🏦",
    layout="wide",
)

st.title("🏦 Caixas e Conciliação Bancária")
st.caption("OFX (Extrato Banco) × XLS (Intermediadora / Rede)")

# ── CSS: simplifica visual dos file_uploaders na sidebar ──
st.markdown("""
<style>
/* Oculta "Drag and drop file here" e "Limit 200MB per file • XLS" */
[data-testid="stFileUploaderDropzoneInstructions"] {
    display: none !important;
}
/* Reduz dropzone — vira só o botão Browse */
[data-testid="stFileUploaderDropzone"] {
    padding: 0.35rem 0.5rem !important;
    min-height: unset !important;
}
/* Botão Browse ocupa largura total */
[data-testid="stFileUploaderDropzone"] button {
    width: 100% !important;
    font-size: 0.78rem !important;
    padding: 0.2rem 0.5rem !important;
}
/* Menos espaço entre uploaders */
[data-testid="stSidebar"] [data-testid="stFileUploader"] {
    margin-bottom: 0.2rem !important;
}
</style>
""", unsafe_allow_html=True)

# Banner de estabelecimento — exibido após o carregamento
_estab_placeholder = st.empty()

# ─────────────────────────────────────────────
# FUNÇÕES DE PARSE
# ─────────────────────────────────────────────

def parse_ofx(file_bytes: bytes) -> pd.DataFrame:
    """Parser manual do OFX. Extrai: data, valor, descrição, fitid."""
    content = file_bytes.decode("latin-1", errors="ignore")
    transactions = []

    pattern = re.compile(r"<STMTTRN>(.*?)</STMTTRN>", re.DOTALL | re.IGNORECASE)
    for match in pattern.finditer(content):
        block = match.group(1)

        def get_field(tag):
            m = re.search(rf"<{tag}>(.*?)(?=<|\Z)", block, re.IGNORECASE | re.DOTALL)
            return m.group(1).strip() if m else ""

        dtposted = get_field("DTPOSTED")
        trnamt   = get_field("TRNAMT")
        memo     = get_field("MEMO") or get_field("NAME")
        fitid    = get_field("FITID")
        trntype  = get_field("TRNTYPE")

        try:
            date = datetime.strptime(dtposted[:8], "%Y%m%d").date()
        except Exception:
            continue
        try:
            valor = float(trnamt.replace(",", "."))
        except Exception:
            continue

        transactions.append({
            "data": date, "valor_ofx": valor,
            "memo": memo, "fitid": fitid, "tipo_lancamento": trntype,
        })

    return pd.DataFrame(transactions)


def detectar_bandeira_tipo(memo: str):
    """Extrai bandeira e tipo (CREDITO/DEBITO) do memo do OFX.
    Sufixos reconhecidos:
      CD / AT → CREDITO  (AT = antecipação de crédito)
      DB      → DEBITO
    MAST é alias de MASTERCARD nos memos da Rede.
    """
    memo_upper = memo.upper()
    bandeiras_map = [
        ("AMERICAN EXPRESS", "AMEX"),
        ("MASTERCARD",       "MASTERCARD"),
        ("MASTER",           "MASTERCARD"),
        ("MAST",             "MASTERCARD"),
        ("HIPERCARD",        "HIPERCARD"),
        ("HIPER",            "HIPERCARD"),
        ("VISA",             "VISA"),
        ("ELO",              "ELO"),
        ("AMEX",             "AMEX"),
        ("CABAL",            "CABAL"),
        ("DINERS",           "DINERS"),
    ]
    bandeira = "OUTROS"
    for token, nome in bandeiras_map:
        if token in memo_upper:
            bandeira = nome
            break

    tipo = "OUTROS"
    if re.search(r"(?<![A-Z])DB(?=\d|$|\s)", memo_upper):
        tipo = "DEBITO"
    elif re.search(r"(?<![A-Z])CD(?=\d|$|\s)|(?<![A-Z])AT(?=\d|$|\s)", memo_upper):
        tipo = "CREDITO"
    elif re.search(r"\bDEBITO\b|\bDÉBITO\b|\bDEBIT\b", memo_upper):
        tipo = "DEBITO"
    elif re.search(r"\bCREDITO\b|\bCRÉDITO\b|\bCREDIT\b", memo_upper):
        tipo = "CREDITO"

    return bandeira, tipo


def proximo_dia_util(data):
    """Retorna o próximo dia útil (seg-sex) a partir de data (inclusive se já for útil)."""
    from datetime import timedelta
    d = data
    while d.weekday() >= 5:   # 5=sábado, 6=domingo
        d += timedelta(days=1)
    return d

def adicionar_dias_uteis(data, n):
    """Soma n dias úteis (seg-sex) a uma data."""
    from datetime import timedelta
    d = data
    adicionados = 0
    while adicionados < n:
        d += timedelta(days=1)
        if d.weekday() < 5:
            adicionados += 1
    return d

def calcular_previsao(data, tipo_norm: str,
                      prazo_debito: int, prazo_credito: int,
                      credito_modo: str) -> str:
    """Calcula data provável de recebimento no banco."""
    from datetime import timedelta
    try:
        if tipo_norm == "DEBITO":
            prev = adicionar_dias_uteis(data, prazo_debito)
        else:  # CREDITO (inclui antecipação AT)
            if credito_modo == "Dias úteis":
                prev = adicionar_dias_uteis(data, prazo_credito)
            else:
                prev = data + timedelta(days=prazo_credito)
                prev = proximo_dia_util(prev)
        return prev.strftime("%d/%m/%Y")
    except Exception:
        return ""


def parse_intermediadora_xls(file) -> pd.DataFrame:
    """Parser para o relatório TSV da intermediadora (.xls)."""
    def to_float(s):
        if pd.isna(s) or str(s).strip() in ("", "-", "nan"): return 0.0
        s = re.sub(r"\(.*?\)", "", str(s).strip())
        nums = re.findall(r"[\d.,]+", s)
        if not nums: return 0.0
        try: return float(nums[0].replace(".", "").replace(",", "."))
        except: return 0.0

    try:
        content = file.read()
        text = content.decode("latin-1", errors="ignore") if isinstance(content, bytes) else content
        df_raw = pd.read_csv(io.StringIO(text), sep="\t", dtype=str, encoding=None)
    except Exception as e:
        raise ValueError(f"Erro ao ler arquivo da intermediadora: {e}")

    df_raw.columns = [c.strip() for c in df_raw.columns]

    df = pd.DataFrame()
    df["data"]          = pd.to_datetime(df_raw["DATA VENDA"], dayfirst=True, errors="coerce").dt.date
    df["bandeira"]      = df_raw["BANDEIRA"].astype(str).str.strip().str.upper()
    df["tipo"]          = df_raw["TRANSACAO"].astype(str).str.strip() if "TRANSACAO" in df_raw.columns else df_raw["TRANSAÇÃO"].astype(str).str.strip()
    df["valor_bruto"]   = df_raw["VALOR BRUTO"].apply(to_float)
    df["valor_liquido"] = df_raw["VALOR LÍQUIDO"].apply(to_float) if "VALOR LÍQUIDO" in df_raw.columns else df_raw["VALOR LIQUIDO"].apply(to_float)
    df["taxa_final"]    = df_raw["TAXA FINAL (R$)"].apply(to_float)
    df["taxa_pct"]      = df_raw["TAXA FINAL (%)"].apply(to_float)
    df["status_adq"]    = df_raw.get("STATUS VENDA", pd.Series([""] * len(df_raw))).astype(str)
    df["cv"]            = df_raw.get("C.V.", pd.Series([""] * len(df_raw))).astype(str)
    df["estabelecimento"] = df_raw.get("ESTABELECIMENTO", pd.Series([""] * len(df_raw))).astype(str).str.strip()
    df["captura"]         = df_raw.get("CAPTURA", pd.Series([""] * len(df_raw))).astype(str).str.strip().str.upper()

    df["bandeira"] = df["bandeira"].replace({
        "MASTER": "MASTERCARD", "MC": "MASTERCARD",
        "AMERICAN EXPRESS": "AMEX", "AX": "AMEX", "VI": "VISA",
    })
    df["tipo_norm"] = df["tipo"].apply(
        lambda x: "CREDITO" if "CRÉD" in x.upper() or "CRED" in x.upper()
                  else ("DEBITO" if "DÉB" in x.upper() or "DEB" in x.upper() else x.upper())
    )

    df = df.dropna(subset=["data"])
    df = df[df["valor_bruto"] > 0].reset_index(drop=True)
    df["idx_transacao"] = df.index
    return df


def agrupar_rede(df_rede: pd.DataFrame) -> pd.DataFrame:
    """Agrupa por data + bandeira + tipo, preservando lista de índices por grupo."""
    group_cols = ["data", "bandeira", "tipo_norm"]
    agg = {}
    for c in ["valor_bruto", "valor_liquido", "taxa_final"]:
        if c in df_rede.columns: agg[c] = "sum"
    agg["idx_transacao"] = lambda x: list(x)

    df_group = df_rede.groupby(group_cols, as_index=False).agg(agg)
    df_group["qtd_transacoes"] = df_group["idx_transacao"].apply(len)
    df_group = df_group.reset_index(drop=True)
    df_group["idx_grupo"] = df_group.index.astype(int)
    return df_group


def conciliar(df_ofx: pd.DataFrame, df_rede_grupo: pd.DataFrame,
              tolerancia_dias: int = 1, tolerancia_valor: float = 0.05) -> pd.DataFrame:
    """Cruza OFX com grupos da intermediadora. Mantém idx_grupo para rastreio."""
    df_ofx = df_ofx.copy()
    bandeira_tipo = df_ofx["memo"].apply(lambda m: pd.Series(detectar_bandeira_tipo(m)))
    df_ofx[["bandeira_ofx", "tipo_ofx"]] = bandeira_tipo

    resultados  = []
    rede_usados = set()
    rede_rows   = df_rede_grupo.reset_index(drop=True)

    for _, row_ofx in df_ofx.iterrows():
        melhor_match = None
        melhor_diff  = float("inf")

        for i_rede, row_rede in rede_rows.iterrows():
            if i_rede in rede_usados: continue
            if (row_ofx["bandeira_ofx"] not in ("OUTROS", "") and
                    row_rede.get("bandeira", "") not in ("OUTROS", "") and
                    row_ofx["bandeira_ofx"] != row_rede.get("bandeira", "")): continue
            if (row_ofx["tipo_ofx"] not in ("OUTROS", "") and
                    row_rede.get("tipo_norm", "") not in ("OUTROS", "") and
                    row_ofx["tipo_ofx"] != row_rede.get("tipo_norm", "")): continue
            diff_dias = abs((row_ofx["data"] - row_rede["data"]).days)
            if diff_dias > tolerancia_dias: continue
            val_ofx  = abs(row_ofx["valor_ofx"])
            val_rede = abs(row_rede.get("valor_liquido", row_rede.get("valor_bruto", 0)))
            if val_rede == 0: continue
            diff_perc = abs(val_ofx - val_rede) / val_rede
            if diff_perc > tolerancia_valor: continue
            score = diff_dias + diff_perc
            if score < melhor_diff:
                melhor_diff  = score
                melhor_match = i_rede

        row_rede_matched  = {}
        idx_grupo_matched = None

        if melhor_match is not None:
            rede_usados.add(melhor_match)
            row_rede_matched  = rede_rows.loc[melhor_match].to_dict()
            idx_grupo_matched = row_rede_matched.get("idx_grupo")
            diff_valor = abs(row_ofx["valor_ofx"]) - abs(row_rede_matched.get("valor_liquido", 0))
            status = "⚠️ Conciliado c/ Divergência" if abs(diff_valor) > 0.01 else "✅ Conciliado"
        else:
            status     = "❌ Não Conciliado (banco)"
            diff_valor = None

        resultados.append({
            "Status":           status,
            "idx_grupo":        idx_grupo_matched,
            "fitid_ofx":        row_ofx.get("fitid", ""),
            "Data OFX":         row_ofx["data"],
            "Valor OFX":        row_ofx["valor_ofx"],
            "Memo OFX":         row_ofx["memo"],
            "Bandeira OFX":     row_ofx["bandeira_ofx"],
            "Tipo OFX":         row_ofx["tipo_ofx"],
            "Data Rede":        row_rede_matched.get("data", ""),
            "Bandeira Rede":    row_rede_matched.get("bandeira", ""),
            "Tipo Rede":        row_rede_matched.get("tipo_norm", ""),
            "Valor Bruto Rede": row_rede_matched.get("valor_bruto", ""),
            "Valor Líq. Rede":  row_rede_matched.get("valor_liquido", ""),
            "Qtd Transações":   row_rede_matched.get("qtd_transacoes", ""),
            "Diferença (R$)":   diff_valor,
        })

    for i_rede, row_rede in rede_rows.iterrows():
        if i_rede not in rede_usados:
            resultados.append({
                "Status":           "❌ Não Conciliado (Rede)",
                "idx_grupo":        row_rede.get("idx_grupo"),
                "fitid_ofx":        "",
                "Data OFX":         "", "Valor OFX": "", "Memo OFX": "",
                "Bandeira OFX":     "", "Tipo OFX": "",
                "Data Rede":        row_rede.get("data", ""),
                "Bandeira Rede":    row_rede.get("bandeira", ""),
                "Tipo Rede":        row_rede.get("tipo_norm", ""),
                "Valor Bruto Rede": row_rede.get("valor_bruto", ""),
                "Valor Líq. Rede":  row_rede.get("valor_liquido", ""),
                "Qtd Transações":   row_rede.get("qtd_transacoes", ""),
                "Diferença (R$)":   None,
            })

    return pd.DataFrame(resultados)


def build_status_transacao(df_rede_orig: pd.DataFrame,
                            df_rede_grupo: pd.DataFrame,
                            df_result: pd.DataFrame,
                            vinculos_manuais: dict) -> pd.DataFrame:
    """Propaga status do grupo (auto + manual) para cada transação individual.
    Vínculos manuais podem ser por idx_grupo (legado) ou por idx_transacao direta (virtual).
    """
    # ── Status automáticos: idx_grupo → status ──
    status_por_grupo = {}
    memo_por_grupo   = {}
    valor_por_grupo  = {}
    for _, row in df_result.iterrows():
        ig = row["idx_grupo"]
        if ig is not None:
            status_por_grupo[ig] = row["Status"]
            memo_por_grupo[ig]   = row.get("Memo OFX", "")
            valor_por_grupo[ig]  = row.get("Valor OFX", "")

    # ── Vínculos manuais: podem ser por idx_grupo ou por idx_transacao ──
    # status_por_transacao tem prioridade sobre status_por_grupo
    status_por_transacao = {}
    memo_por_transacao   = {}
    valor_por_transacao  = {}

    for chave, info in vinculos_manuais.items():
        if info.get("virtual"):
            # Vínculo direto por transação individual
            for idx_t in info.get("idx_transacoes", []):
                status_por_transacao[idx_t] = info["status"]
                memo_por_transacao[idx_t]   = info.get("memo_ofx", "")
                valor_por_transacao[idx_t]  = info.get("valor_ofx", "")
        else:
            # Vínculo por grupo (legado)
            try:
                ig = int(chave)
                status_por_grupo[ig] = info["status"]
                memo_por_grupo[ig]   = info.get("memo_ofx", "")
                valor_por_grupo[ig]  = info.get("valor_ofx", "")
            except (ValueError, TypeError):
                pass

    # Mapa idx_transacao → idx_grupo
    mapa_idx = {}
    for _, row in df_rede_grupo.iterrows():
        for idx_t in row["idx_transacao"]:
            mapa_idx[idx_t] = row["idx_grupo"]

    df = df_rede_orig.copy()
    df["idx_grupo"] = df["idx_transacao"].map(mapa_idx)

    def get_status(row):
        idx_t = row["idx_transacao"]
        if idx_t in status_por_transacao:
            return status_por_transacao[idx_t]
        ig = row["idx_grupo"]
        return status_por_grupo.get(ig, "❌ Não Conciliado (Rede)")

    def get_memo(row):
        idx_t = row["idx_transacao"]
        if idx_t in memo_por_transacao:
            return memo_por_transacao[idx_t]
        ig = row["idx_grupo"]
        return memo_por_grupo.get(ig, "")

    def get_valor(row):
        idx_t = row["idx_transacao"]
        if idx_t in valor_por_transacao:
            return valor_por_transacao[idx_t]
        ig = row["idx_grupo"]
        return valor_por_grupo.get(ig, "")

    df["Status Conciliação"] = df.apply(get_status, axis=1)
    df["Memo OFX Vinculado"]  = df.apply(get_memo,   axis=1)
    df["Valor OFX Vinculado"] = df.apply(get_valor,  axis=1)

    return df


def exportar_excel(df_result: pd.DataFrame,
                   df_detalhe_status: pd.DataFrame,
                   df_rede_grupo: pd.DataFrame,
                   vinculos_manuais: dict) -> bytes:
    output = io.BytesIO()

    df_result_exp = df_result.drop(columns=["idx_grupo", "fitid_ofx"], errors="ignore")
    df_grupo_exp  = df_rede_grupo.drop(columns=["idx_transacao", "idx_grupo"], errors="ignore")
    cols_det      = [c for c in df_detalhe_status.columns if c not in ("idx_transacao", "idx_grupo")]
    df_det_exp    = df_detalhe_status[cols_det]

    # Aba de vínculos manuais
    registros_manual = []
    for ig, info in vinculos_manuais.items():
        registros_manual.append({
            "Status":          info["status"],
            "Memo OFX":        info.get("memo_ofx", ""),
            "Valor OFX":       info.get("valor_ofx", ""),
            "Data OFX":        str(info.get("data_ofx", "")),
            "Diferença (R$)":  info.get("diff_valor", ""),
            "Observação":      info.get("observacao", ""),
        })
    df_manual = pd.DataFrame(registros_manual) if registros_manual else pd.DataFrame()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df_result_exp.to_excel(writer, sheet_name="Conciliação",        index=False)
        df_grupo_exp.to_excel(writer,  sheet_name="Grupos",             index=False)
        df_det_exp.to_excel(writer,    sheet_name="Detalhe Transações", index=False)
        if not df_manual.empty:
            df_manual.to_excel(writer, sheet_name="Vínculos Manuais",   index=False)

        wb    = writer.book
        fmt_h    = wb.add_format({"bold": True, "bg_color": "#1F3864", "font_color": "white", "border": 1})
        fmt_ok   = wb.add_format({"bg_color": "#C6EFCE"})
        fmt_warn = wb.add_format({"bg_color": "#FFEB9C"})
        fmt_man  = wb.add_format({"bg_color": "#DDEBF7"})
        fmt_err  = wb.add_format({"bg_color": "#FFC7CE"})

        def aplicar_formato(ws, df_exp, status_col):
            for col_num, col_name in enumerate(df_exp.columns):
                ws.write(0, col_num, col_name, fmt_h)
                ws.set_column(col_num, col_num, 22)
            if status_col in df_exp.columns:
                for row_num in range(1, len(df_exp) + 1):
                    s = str(df_exp.iloc[row_num - 1][status_col])
                    if "✅" in s:    fmt = fmt_ok
                    elif "⚠️" in s: fmt = fmt_warn
                    elif "🔗" in s: fmt = fmt_man
                    else:            fmt = fmt_err
                    ws.set_row(row_num, None, fmt)

        aplicar_formato(writer.sheets["Conciliação"],        df_result_exp, "Status")
        aplicar_formato(writer.sheets["Detalhe Transações"], df_det_exp,    "Status Conciliação")

    return output.getvalue()


# ─────────────────────────────────────────────
# FUNÇÕES — ESTABELECIMENTOS
# ─────────────────────────────────────────────

def parse_estabelecimentos(file) -> pd.DataFrame:
    """Lê Lista_Estabelecimentos.xlsx → DataFrame com Fantasia, Abreviação, CNPJ, ESTABELECIMENTO, ACCTID."""
    df = pd.read_excel(file)
    df.columns = [c.strip() for c in df.columns]

    def int_str(v):
        """Converte numérico (int ou float) para string sem '.0'.
        Ex: 7197997393.0 → '7197997393'   NaN → ''"""
        try:
            if pd.isna(v): return ""
            return str(int(float(str(v)))).strip()
        except:
            return str(v).strip()

    df["ACCTID"]          = df["ACCTID"].apply(int_str)
    df["ESTABELECIMENTO"] = df["ESTABELECIMENTO"].apply(int_str)
    # Garante coluna Abreviação mesmo se não existir no arquivo
    if "Abreviação" not in df.columns:
        df["Abreviação"] = ""
    df["Abreviação"] = df["Abreviação"].astype(str).str.strip()
    return df


def acctid_do_ofx(file_bytes: bytes) -> str:
    """Extrai o ACCTID do bloco <ACCTID> do OFX.
    Suporta formatos SGML (valor inline) e XML (<ACCTID>valor</ACCTID>).
    """
    content = file_bytes.decode("latin-1", errors="ignore")
    # Tenta formato SGML: <ACCTID>7197997393  (valor até próxima tag, newline ou fim)
    m = re.search(r"<ACCTID>\s*([^\s<]+)", content, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    # Tenta formato XML: <ACCTID>7197997393</ACCTID>
    m = re.search(r"<ACCTID>(.*?)</ACCTID>", content, re.IGNORECASE | re.DOTALL)
    return m.group(1).strip() if m else ""


def nome_filial_do_arquivo(filename: str) -> str:
    """
    Extrai nome da filial do nome do arquivo de caixa.
    Regra: texto após o primeiro '.' no stem, sem o prefixo 'drops_'.
    Ex: pagamentos_listagem_2026_1.drops_brasilia.xlsx → BRASÍLIA
    """
    import os
    stem = os.path.splitext(filename)[0]          # remove extensão
    parts = stem.split(".", 1)
    if len(parts) < 2:
        return stem.upper()
    filial = parts[1].lower()
    filial = re.sub(r"^drops_", "", filial)        # remove prefixo drops_
    filial = filial.replace("_", " ").strip().upper()
    return filial


# ─────────────────────────────────────────────
# FUNÇÕES — CAIXA
# ─────────────────────────────────────────────

def parse_caixa(file) -> pd.DataFrame:
    """Lê relatório de caixa (xlsx com header na linha 1)."""
    df = pd.read_excel(file, header=1)
    df.columns = [str(c).strip() for c in df.columns]

    # Normaliza tipo de pagamento
    def norm_forma(f):
        """Normaliza Forma Pagamento para CREDITO / DEBITO / PIX / <MAIÚSCULO>.
        Cobre variações entre unidades:
          Crédito, Crédito 3x, Crédito 2x, Credito, CREDITO, Cartão Crédito, Cartao Credito → CREDITO
          Débito, Debito, Cartão Debito, Cartao Débito, DEBITO                               → DEBITO
          PIX, Pix                                                                            → PIX
        """
        fu = str(f).strip().upper()
        # Remove acentos para comparação robusta
        import unicodedata
        fu_norm = ''.join(
            c for c in unicodedata.normalize('NFD', fu)
            if unicodedata.category(c) != 'Mn'
        )
        if "CREDITO" in fu_norm or "CRÉDITO" in fu:  return "CREDITO"
        if "DEBITO"  in fu_norm or "DÉBITO"  in fu:  return "DEBITO"
        if "PIX" in fu_norm:                          return "PIX"
        return fu

    df["forma_norm"] = df["Forma Pagamento"].apply(norm_forma)

    # Normaliza bandeira
    df["bandeira_norm"] = df["Bandeira"].astype(str).str.upper().str.strip()
    df["bandeira_norm"] = df["bandeira_norm"].replace({
        "MASTER": "MASTERCARD", "AMERICAN EXPRESS": "AMEX",
        "NAN": "", "NONE": ""
    })

    # Data como datetime
    df["data_hora"] = pd.to_datetime(df["Data Pagamento"], errors="coerce")
    df["data"]      = df["data_hora"].dt.date

    # Valor numérico
    def to_float(v):
        try: return float(str(v).replace(",", ".").strip())
        except: return 0.0
    df["valor"]     = df["Valor"].apply(to_float)

    # AutExtRef como string limpa
    df["AutExtRef"] = df["AutExtRef"].astype(str).str.strip().replace("nan", "")

    return df.dropna(subset=["data_hora"]).reset_index(drop=True)


def conciliar_caixa_rede(df_caixa: pd.DataFrame,
                          df_rede: pd.DataFrame) -> pd.DataFrame:
    """
    Cruza caixa (cartões) com intermediadora via merge vetorizado — O(n) em vez de O(n²).

    Etapa 1 : C.V. exato (AutExtRef = C.V.) — qualquer data
    Etapa 1.5: E-commerce — valor + bandeira, até 90 dias
    Etapa 2 : valor + bandeira + data ≤ 1 dia
    Resto   : ❌ → vinculação manual
    """
    import numpy as np

    df_cart  = df_caixa[df_caixa["forma_norm"].isin(["CREDITO", "DEBITO"])].copy().reset_index(drop=True)
    df_rede2 = df_rede.copy().reset_index(drop=True)
    df_rede2["cv_str"]  = df_rede2["cv"].astype(str).str.strip()
    df_rede2["captura"] = df_rede2.get("captura", pd.Series([""] * len(df_rede2))).astype(str).str.upper()

    # Índice permanente para rastrear quais registros da Rede já foram usados
    df_rede2["_rede_idx"] = df_rede2.index
    df_cart["_cart_idx"]  = df_cart.index

    # Colunas de saída padrão
    COLS_REDE = ["_rede_idx", "cv", "data", "bandeira", "valor_bruto", "captura"]
    vazia = {c: "" for c in ["C.V. Rede","Data Rede","Bandeira Rede","Valor Rede","Captura"]}

    resultado_rows = []
    rede_usados = set()

    def primeiro_match(candidatos):
        """Retorna (rede_idx, row) do primeiro candidato não usado, ou None."""
        for _, r in candidatos.iterrows():
            if r["_rede_idx"] not in rede_usados:
                return r
        return None

    # ── Pré-build de índices para lookup rápido ──────────────────────────
    # CV → lista de rede_idx
    cv_index = {}
    for r_idx, r in df_rede2.iterrows():
        cv = r["cv_str"]
        if cv and cv not in ("nan",""):
            cv_index.setdefault(cv, []).append(r_idx)

    # (valor_centavos, bandeira) → lista de (rede_idx, data)  — para etapas 1.5 e 2
    val_band_index = {}
    for r_idx, r in df_rede2.iterrows():
        key = (round(r["valor_bruto"], 2), r["bandeira"])
        val_band_index.setdefault(key, []).append(r_idx)

    # ── Loop por registro do caixa ───────────────────────────────────────
    for _, rc in df_cart.iterrows():
        status        = "❌ Não encontrado na Rede"
        match_cv      = ""
        match_data    = ""
        match_val     = ""
        match_band    = ""
        match_captura = ""

        # Etapa 1: C.V. exato
        if rc["AutExtRef"] and rc["AutExtRef"] not in ("", "nan"):
            for r_idx in cv_index.get(rc["AutExtRef"], []):
                if r_idx in rede_usados:
                    continue
                r = df_rede2.iloc[r_idx]
                diff_val = abs(rc["valor"] - r["valor_bruto"])
                rede_usados.add(r_idx)
                status        = "✅ Conciliado (C.V.)" if diff_val < 0.02 else "⚠️ C.V. ok, valor diverge"
                match_cv      = r["cv"]
                match_data    = str(r["data"])
                match_val     = r["valor_bruto"]
                match_band    = r["bandeira"]
                match_captura = r["captura"]
                break

        # Etapa 1.5: E-commerce, até 90 dias
        if "❌" in status:
            data_cx = rc["data"]
            key = (round(rc["valor"], 2), rc["bandeira_norm"])
            for r_idx in val_band_index.get(key, []):
                if r_idx in rede_usados:
                    continue
                r = df_rede2.iloc[r_idx]
                if r["captura"] != "E-COMMERCE":
                    continue
                try:
                    diff_dias = abs((data_cx - r["data"]).days)
                except Exception:
                    diff_dias = 999
                if diff_dias > 90:
                    continue
                rede_usados.add(r_idx)
                status        = "✅ E-commerce (reserva)"
                match_cv      = r["cv"]
                match_data    = str(r["data"])
                match_val     = r["valor_bruto"]
                match_band    = r["bandeira"]
                match_captura = r["captura"]
                break

        # Etapa 2: valor + bandeira + data ≤ 1 dia
        if "❌" in status:
            data_cx = rc["data"]
            key = (round(rc["valor"], 2), rc["bandeira_norm"])
            for r_idx in val_band_index.get(key, []):
                if r_idx in rede_usados:
                    continue
                r = df_rede2.iloc[r_idx]
                try:
                    diff_dias = abs((data_cx - r["data"]).days)
                except Exception:
                    diff_dias = 999
                if diff_dias > 1:
                    continue
                rede_usados.add(r_idx)
                status        = "✅ Conciliado (valor+data)"
                match_cv      = r["cv"]
                match_data    = str(r["data"])
                match_val     = r["valor_bruto"]
                match_band    = r["bandeira"]
                match_captura = r["captura"]
                break

        resultado_rows.append({
            "Status":        status,
            "Caixa":         rc["Caixa"],
            "Forma":         rc["Forma Pagamento"],
            "forma_norm":    rc["forma_norm"],
            "Bandeira Cx":   rc["bandeira_norm"],
            "Data/Hora":     rc["data_hora"].strftime("%d/%m/%Y %H:%M") if pd.notna(rc["data_hora"]) else "",
            "Valor Caixa":   rc["valor"],
            "AutExtRef":     rc["AutExtRef"],
            "C.V. Rede":     match_cv,
            "Data Rede":     match_data,
            "Bandeira Rede": match_band,
            "Valor Rede":    match_val,
            "Captura":       match_captura,
        })

    # Registros da Rede sem par no caixa
    for idx2, r2 in df_rede2.iterrows():
        if idx2 not in rede_usados:
            resultado_rows.append({
                "Status":        "❌ Não encontrado no Caixa",
                "Caixa":         "",
                "Forma":         r2.get("tipo_norm", ""),
                "forma_norm":    r2.get("tipo_norm", ""),
                "Bandeira Cx":   "",
                "Data/Hora":     "",
                "Valor Caixa":   "",
                "AutExtRef":     "",
                "C.V. Rede":     r2["cv"],
                "Data Rede":     str(r2["data"]),
                "Bandeira Rede": r2["bandeira"],
                "Valor Rede":    r2["valor_bruto"],
                "Captura":       str(r2.get("captura", "")),
            })

    return pd.DataFrame(resultado_rows)


# ─────────────────────────────────────────────
# FUNÇÕES — PIX POS
# ─────────────────────────────────────────────

def acctid_do_nome_arquivo_pix(filename: str) -> str:
    """
    Extrai ACCTID do nome do arquivo PIX POS.
    Formato: NOME_AGENCIA_CONTA_....xls
    Ex: SP_HOTELARIA_SPE_LTDA_7197_997393_... → ACCTID = 7197997393
    Remove o traço do dígito verificador: 99739-3 → 997393
    """
    import os
    stem = os.path.splitext(filename)[0]
    # Busca dois segmentos numéricos consecutivos (agência + conta)
    partes = stem.split("_")
    for i in range(len(partes) - 1):
        ag = partes[i].strip()
        ct = partes[i+1].strip().replace("-","")
        if ag.isdigit() and ct.isdigit() and len(ag) <= 6 and len(ct) <= 8:
            return ag + ct
    return ""


def parse_pix_pos(file, df_estab: pd.DataFrame = None) -> tuple:
    """
    Lê relatório PIX POS em dois formatos:

    XLS/XLSX (Sicredi padrão por filial):
      Cabeçalho com Nome/Agência/Conta, dados a partir da linha "identificador".

    CSV (Google Sheets consolidado — múltiplas filiais):
      Colunas: Data Completa, Filial, Caixa, <filiais...>, identificador,
               pagador efetivo, cpf/cnpj, vencimento, pago em,
               valor emitido (R$), valor pago (R$), tarifa (R$)
      Identificação de filial via coluna "Filial" cruzada com df_estab["Abreviação"].

    Retorna (df_transacoes, metadata).
    """
    import csv as _csv

    raw = file.read()

    def br_float(s):
        try:
            return float(str(s).replace(".", "").replace(",", ".").strip())
        except:
            return 0.0

    def parse_data(s):
        s = str(s).strip()
        try:
            return pd.to_datetime(s[:10], format="%d/%m/%Y").date()
        except:
            return None

    # ── Formato CSV (Google Sheets) ──────────────────────────────────────
    is_csv = (raw[:2] not in (b"\xd0\xcf", b"\xD0\xCF")
              and raw[:4] not in (b"\xd0\xcf\x11\xe0", b"PK\x03\x04"))
    if is_csv:
        text   = raw.decode("utf-8", errors="replace")
        reader = _csv.reader(text.splitlines())
        header = [h.strip() for h in next(reader)]

        # Índices das colunas relevantes
        def idx_of(*names):
            for n in names:
                if n in header: return header.index(n)
            return None

        idx_filial  = idx_of("Abrev", "Filial")  # "Abrev" é o padrão; "Filial" como fallback
        idx_ident   = idx_of("identificador")
        idx_pagador = idx_of("pagador efetivo")
        idx_cpf     = idx_of("cpf/cnpj")
        idx_venc    = idx_of("vencimento ou expiração")
        idx_pago_em = idx_of("pago em")
        idx_emitido = idx_of("valor emitido (R$)")
        idx_pago    = idx_of("valor pago (R$)")
        idx_tarifa  = idx_of("tarifa (R$)")

        # Mapa abrev → Fantasia
        abrev_map = {}
        if df_estab is not None and not df_estab.empty and "Abreviação" in df_estab.columns:
            for _, er in df_estab.iterrows():
                abrev_map[str(er["Abreviação"]).strip().upper()] = er["Fantasia"]

        def get(row, idx):
            return row[idx].strip() if idx is not None and idx < len(row) else ""

        rows = []
        for row in reader:
            if not row or not row[0].strip():
                continue
            ident = get(row, idx_ident)
            if not ident or not ident.startswith("RESD"):
                continue
            abrev    = get(row, idx_filial).upper()
            fantasia = abrev_map.get(abrev, abrev)
            rows.append({
                "identificador":  ident,
                "pagador":        get(row, idx_pagador),
                "cpf_cnpj":       get(row, idx_cpf),
                "vencimento":     get(row, idx_venc),
                "pago_em":        get(row, idx_pago_em),
                "valor_emitido":  get(row, idx_emitido),
                "valor_pago":     get(row, idx_pago),
                "tarifa":         get(row, idx_tarifa),
                "abrev":          abrev,
                "filial":         fantasia,
            })

        if not rows:
            return pd.DataFrame(), {"formato": "csv_sheets"}

        df = pd.DataFrame(rows)
        df["valor_pago_num"]    = df["valor_pago"].apply(br_float)
        df["valor_emitido_num"] = df["valor_emitido"].apply(br_float)
        df["tarifa_num"]        = df["tarifa"].apply(br_float)
        df["data_pago"]         = df["pago_em"].apply(parse_data)
        df = df.dropna(subset=["data_pago"]).reset_index(drop=True)
        df["idx_pix"] = df.index

        filiais = df["abrev"].unique().tolist()
        meta = {
            "formato":            "csv_sheets",
            "total_filiais":      len(filiais),
            "filiais":            ", ".join(sorted(filiais)),
            "periodo_inicio":     str(df["data_pago"].min()),
            "periodo_fim":        str(df["data_pago"].max()),
            "total_recebimentos": f"R$ {df['valor_pago_num'].sum():,.2f}",
        }
        return df, meta

    # ── Formato XLS/XLSX (Sicredi padrão por filial) ─────────────────────
    if raw[:2] in (b"\xd0\xcf", b"\xD0\xCF") or raw[:4] == b"\xd0\xcf\x11\xe0":
        import xlrd
        wb_xl = xlrd.open_workbook(file_contents=raw)
        ws_xl = wb_xl.sheet_by_index(0)
        all_rows_raw = [
            [str(ws_xl.cell_value(i, j)).strip() for j in range(ws_xl.ncols)]
            for i in range(ws_xl.nrows)
        ]
    else:
        import openpyxl, io as _io
        wb_xl = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
        ws_xl = wb_xl.active
        all_rows_raw = [
            [str(cell.value).strip() if cell.value is not None else ""
             for cell in row]
            for row in ws_xl.iter_rows()
        ]

    meta = {}
    header_row = None
    rows = []

    for i, full_row in enumerate(all_rows_raw):
        row_vals = [v for v in full_row if v not in ("", "nan", "None")]
        if not row_vals:
            continue
        label = row_vals[0].lower()
        if "nome" in label and len(row_vals) > 1:
            meta["nome"] = row_vals[1]
        elif "agência" in label or "agencia" in label:
            meta["agencia"] = row_vals[1] if len(row_vals) > 1 else ""
        elif "conta corrente" in label:
            meta["conta"] = row_vals[1] if len(row_vals) > 1 else ""
        elif "período" in label or "periodo" in label:
            meta["periodo"] = row_vals[1] if len(row_vals) > 1 else ""
        elif row_vals[0].lower() == "identificador":
            header_row = i
        elif header_row is not None and i > header_row:
            if len(full_row) >= 6 and full_row[0]:
                rows.append(full_row)

    ag  = meta.get("agencia", "").strip()
    ct  = meta.get("conta", "").strip().replace("-","")
    meta["acctid"]  = ag + ct if ag and ct else ""
    meta["formato"] = "xls_sicredi"

    if not rows:
        return pd.DataFrame(), meta

    df = pd.DataFrame(rows, columns=[
        "identificador", "pagador", "cpf_cnpj",
        "vencimento", "pago_em", "valor_emitido", "valor_pago", "tarifa",
        *[f"_extra{k}" for k in range(max(0, len(rows[0]) - 8))]
    ])

    df["valor_pago_num"]    = df["valor_pago"].apply(br_float)
    df["valor_emitido_num"] = df["valor_emitido"].apply(br_float)
    df["tarifa_num"]        = df["tarifa"].apply(br_float)
    df["data_pago"]         = df["pago_em"].apply(parse_data)
    df = df.dropna(subset=["data_pago"]).reset_index(drop=True)
    df["idx_pix"] = df.index
    return df, meta


def conciliar_pix_pos_ofx(df_pix: pd.DataFrame,
                            df_ofx: pd.DataFrame,
                            tolerancia_dias: int = 2) -> pd.DataFrame:
    """
    Cruza PIX POS com lançamentos OFX que contêm PIX no memo.
    Chave: valor_pago_num (bruto) + data_pago ± tolerancia_dias.
    """
    # Filtra OFX com PIX no memo (créditos)
    df_ofx_pix = df_ofx[
        df_ofx["memo"].str.upper().str.contains("PIX", na=False) &
        (df_ofx["valor_ofx"] > 0)
    ].copy().reset_index(drop=True)
    df_ofx_pix["_ofx_idx"] = df_ofx_pix.index

    # Índice de lookup: (valor_arredondado) → [ofx_idx]
    val_index = {}
    for i, r in df_ofx_pix.iterrows():
        key = round(r["valor_ofx"], 2)
        val_index.setdefault(key, []).append(i)

    ofx_usados = set()
    resultado  = []

    for _, rp in df_pix.iterrows():
        status     = "❌ Não encontrado no OFX"
        match_data = ""
        match_memo = ""
        match_fitid= ""
        match_val  = ""

        key = round(rp["valor_pago_num"], 2)
        for ofx_idx in val_index.get(key, []):
            if ofx_idx in ofx_usados:
                continue
            ro = df_ofx_pix.iloc[ofx_idx]
            try:
                diff = abs((rp["data_pago"] - ro["data"]).days)
            except:
                diff = 999
            if diff > tolerancia_dias:
                continue
            ofx_usados.add(ofx_idx)
            status      = "✅ Conciliado"
            match_data  = str(ro["data"])
            match_memo  = ro["memo"]
            match_fitid = ro.get("fitid", "")
            match_val   = ro["valor_ofx"]
            break

        resultado.append({
            "Status":       status,
            "Data Pago":    rp["data_pago"].strftime("%d/%m/%Y") if rp["data_pago"] else "",
            "Identificador":rp["identificador"],
            "Pagador":      rp["pagador"],
            "Valor Pago":   rp["valor_pago_num"],
            "Tarifa":       rp["tarifa_num"],
            "Data OFX":     match_data,
            "Memo OFX":     match_memo,
            "Valor OFX":    match_val,
            "FITID OFX":    match_fitid,
            "idx_pix":      int(rp["idx_pix"]),
        })

    # OFX PIX sem par no relatório POS
    for ofx_idx, ro in df_ofx_pix.iterrows():
        if ofx_idx not in ofx_usados:
            resultado.append({
                "Status":       "❌ Não encontrado no POS",
                "Data Pago":    "",
                "Identificador":"",
                "Pagador":      "",
                "Valor Pago":   "",
                "Tarifa":       "",
                "Data OFX":     str(ro["data"]),
                "Memo OFX":     ro["memo"],
                "Valor OFX":    ro["valor_ofx"],
                "FITID OFX":    ro.get("fitid", ""),
                "idx_pix":      -1,
            })

    return pd.DataFrame(resultado), df_ofx_pix


# ─────────────────────────────────────────────
# FUNÇÕES — PIX TEF
# ─────────────────────────────────────────────

def parse_pix_tef(file, df_estab: pd.DataFrame = None) -> tuple:
    """
    Lê relatório PIX TEF em dois formatos:

    XLSX (Sicredi padrão):
      Cabeçalho com Associado/Cooperativa/Conta, dados a partir da linha "Natureza da transação".

    CSV (Google Sheets consolidado — múltiplas filiais):
      Colunas fixas: Data Completa, Abrev, Caixa, <filiais...>, Natureza, Tipo, Nome, Data, Hora, Valor, ID...
      Identificação de filial via coluna "Abrev" cruzada com df_estab["Abreviação"].

    Retorna (df_transacoes, metadata).
    Filtra somente CRÉDITO / RECEBIMENTO.
    """
    import openpyxl, io as _io, csv as _csv

    raw = file.read()

    def br_float(s):
        try:
            return float(str(s).replace("R$","").replace(".","").replace(",",".").strip())
        except:
            return 0.0

    def parse_data(s):
        try:
            return pd.to_datetime(str(s).strip()[:10], format="%d/%m/%Y").date()
        except:
            return None

    # ── Formato CSV (Google Sheets) ──────────────────────────────────────
    is_csv = raw[:4] != b"PK\x03\x04" and raw[:4] != b"\xd0\xcf\x11\xe0"
    if is_csv:
        text   = raw.decode("utf-8", errors="replace")
        reader = _csv.reader(text.splitlines())
        header = next(reader)
        header = [h.strip() for h in header]

        # Índices das colunas relevantes
        idx_abrev    = header.index("Abrev")        if "Abrev"                 in header else None
        idx_nat      = header.index("Natureza da transação") if "Natureza da transação" in header else None
        idx_tipo     = header.index("Tipo da transação")     if "Tipo da transação"     in header else None
        idx_nome     = header.index("Nome")          if "Nome"                  in header else None
        idx_data     = header.index("Data")          if "Data"                  in header else None
        idx_valor    = header.index("Valor")         if "Valor"                 in header else None
        idx_id       = header.index("ID")            if "ID"                    in header else None
        idx_inst     = header.index("Instituição")   if "Instituição"           in header else None
        idx_chave    = header.index("Chave")         if "Chave"                 in header else None
        idx_conc     = header.index("ID Conciliador")if "ID Conciliador"        in header else None

        # Mapa abrev → Fantasia (via df_estab se disponível)
        abrev_map = {}
        if df_estab is not None and not df_estab.empty and "Abreviação" in df_estab.columns:
            for _, er in df_estab.iterrows():
                abrev_map[str(er["Abreviação"]).strip().upper()] = er["Fantasia"]

        def get(row, idx):
            return row[idx].strip() if idx is not None and idx < len(row) else ""

        rows = []
        for row in reader:
            if not row or not row[0].strip():
                continue
            nat  = get(row, idx_nat).upper()
            tipo = get(row, idx_tipo).upper()
            if nat != "CRÉDITO" or tipo != "RECEBIMENTO":
                continue
            abrev = get(row, idx_abrev).upper()
            if abrev in ("#N/A", ""):
                continue
            fantasia = abrev_map.get(abrev, abrev)
            rows.append({
                "natureza":       nat,
                "tipo":           tipo,
                "nome":           get(row, idx_nome),
                "data_str":       get(row, idx_data),
                "valor_str":      get(row, idx_valor),
                "end_to_end_id":  get(row, idx_id),
                "instituicao":    get(row, idx_inst),
                "chave":          get(row, idx_chave),
                "id_conciliador": get(row, idx_conc),
                "abrev":          abrev,
                "filial":         fantasia,
            })

        if not rows:
            return pd.DataFrame(), {"formato": "csv_sheets", "total_filiais": 0}

        df = pd.DataFrame(rows)
        df["valor_num"] = df["valor_str"].apply(br_float)
        df["data_pago"] = df["data_str"].apply(parse_data)
        df = df.dropna(subset=["data_pago"]).reset_index(drop=True)
        df["idx_tef"] = df.index

        filiais = df["abrev"].unique().tolist()
        meta = {
            "formato":        "csv_sheets",
            "total_filiais":  len(filiais),
            "filiais":        ", ".join(sorted(filiais)),
            "periodo_inicio": str(df["data_pago"].min()) if not df.empty else "",
            "periodo_fim":    str(df["data_pago"].max()) if not df.empty else "",
            "total_recebimentos": f"R$ {df['valor_num'].sum():,.2f}",
        }
        return df, meta

    # ── Formato XLSX (Sicredi padrão por filial) ─────────────────────────
    if raw[:4] == b"PK\x03\x04":
        wb  = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
        ws  = wb.active
        all_rows = [
            [str(c.value).strip() if c.value is not None else "" for c in row]
            for row in ws.iter_rows()
        ]
    else:
        # CSV simples (outro formato)
        text = raw.decode("utf-8", errors="replace")
        reader2 = _csv.reader(text.splitlines())
        all_rows = [[str(v).strip() for v in row] for row in reader2]

    meta       = {}
    header_row = None
    rows       = []

    for i, full_row in enumerate(all_rows):
        vals = [v for v in full_row if v not in ("", "nan", "None")]
        if not vals:
            continue
        label = vals[0].lower()
        if "associado" in label and len(vals) > 1:
            meta["associado"] = vals[1]
        elif "data inicio" in label or "data início" in label:
            meta["periodo_inicio"] = vals[1] if len(vals) > 1 else ""
        elif "data fim" in label:
            meta["periodo_fim"] = vals[1] if len(vals) > 1 else ""
        elif "total recebimentos" in label:
            meta["total_recebimentos"] = vals[1] if len(vals) > 1 else ""
        elif vals[0].lower() == "natureza da transação":
            header_row = i
        elif header_row is not None and i > header_row:
            if len(full_row) >= 6 and full_row[0]:
                rows.append(full_row)

    if not rows:
        return pd.DataFrame(), meta

    COLS = ["natureza","tipo","nome","data_str","hora","valor_str",
            "end_to_end_id","origem","tipo_iniciacao","instituicao",
            "descricao","chave","id_conciliador"]
    df = pd.DataFrame(rows, columns=COLS[:len(rows[0])] + [f"_x{k}" for k in range(max(0, len(rows[0])-len(COLS)))])

    df = df[df["natureza"].str.upper() == "CRÉDITO"].copy()
    df = df[df["tipo"].str.upper() == "RECEBIMENTO"].copy()

    df["valor_num"] = df["valor_str"].apply(br_float)
    df["data_pago"] = df["data_str"].apply(parse_data)
    df = df.dropna(subset=["data_pago"]).reset_index(drop=True)
    df["idx_tef"]   = df.index
    meta["formato"] = "xlsx_sicredi"
    return df, meta


def conciliar_pix_unificado(df_pos: pd.DataFrame,
                              df_tef: pd.DataFrame,
                              df_ofx: pd.DataFrame,
                              tolerancia_dias: int = 2) -> tuple:
    """
    Concilia POS + TEF contra OFX PIX em uma passagem única.
    Retorna (df_resultado, df_ofx_pix_usado).
    Coluna Origem: 'POS' ou 'TEF'.
    """
    df_ofx_pix = df_ofx[
        df_ofx["memo"].str.upper().str.contains("PIX", na=False) &
        (df_ofx["valor_ofx"] > 0)
    ].copy().reset_index(drop=True)
    df_ofx_pix["_ofx_idx"] = df_ofx_pix.index

    # Índice hash: valor → [ofx_idx]
    val_index = {}
    for i, r in df_ofx_pix.iterrows():
        val_index.setdefault(round(r["valor_ofx"], 2), []).append(i)

    ofx_usados = set()
    resultado  = []

    def tenta_match(valor_num, data_pago):
        """Retorna (ofx_idx, row_ofx) ou (None, None)."""
        for oi in val_index.get(round(valor_num, 2), []):
            if oi in ofx_usados:
                continue
            ro = df_ofx_pix.iloc[oi]
            try:
                diff = abs((data_pago - ro["data"]).days)
            except:
                diff = 999
            if diff <= tolerancia_dias:
                return oi, ro
        return None, None

    # POS
    if not df_pos.empty:
        for _, rp in df_pos.iterrows():
            oi, ro = tenta_match(rp["valor_pago_num"], rp["data_pago"])
            if oi is not None:
                ofx_usados.add(oi)
                status = "✅ Conciliado"
                md, mm, mf, mv = str(ro["data"]), ro["memo"], ro.get("fitid",""), ro["valor_ofx"]
            else:
                status = "❌ Não encontrado no OFX"
                md, mm, mf, mv = "", "", "", ""
            resultado.append({
                "Status":       status,
                "Origem":       "POS",
                "Filial":       rp.get("filial", rp.get("abrev", "")),
                "Data Pago":    rp["data_pago"].strftime("%d/%m/%Y") if rp["data_pago"] else "",
                "Identificador":rp.get("identificador",""),
                "Pagador":      rp.get("pagador",""),
                "Valor Pago":   rp["valor_pago_num"],
                "Tarifa":       rp.get("tarifa_num", 0),
                "Data OFX":     md, "Memo OFX": mm, "Valor OFX": mv, "FITID OFX": mf,
                "_src_idx":     int(rp["idx_pix"]),
            })

    # TEF
    if not df_tef.empty:
        for _, rt in df_tef.iterrows():
            oi, ro = tenta_match(rt["valor_num"], rt["data_pago"])
            if oi is not None:
                ofx_usados.add(oi)
                status = "✅ Conciliado"
                md, mm, mf, mv = str(ro["data"]), ro["memo"], ro.get("fitid",""), ro["valor_ofx"]
            else:
                status = "❌ Não encontrado no OFX"
                md, mm, mf, mv = "", "", "", ""
            resultado.append({
                "Status":       status,
                "Origem":       "TEF",
                "Filial":       rt.get("filial", rt.get("abrev", "")),
                "Data Pago":    rt["data_pago"].strftime("%d/%m/%Y") if rt["data_pago"] else "",
                "Identificador":rt.get("end_to_end_id",""),
                "Pagador":      rt.get("nome",""),
                "Valor Pago":   rt["valor_num"],
                "Tarifa":       0,
                "Data OFX":     md, "Memo OFX": mm, "Valor OFX": mv, "FITID OFX": mf,
                "_src_idx":     int(rt["idx_tef"]),
            })

    # OFX sem par
    for oi, ro in df_ofx_pix.iterrows():
        if oi not in ofx_usados:
            resultado.append({
                "Status":       "❌ Não encontrado no POS/TEF",
                "Origem":       "OFX",
                "Filial":       "",
                "Data Pago":    "",
                "Identificador":"",
                "Pagador":      "",
                "Valor Pago":   "",
                "Tarifa":       "",
                "Data OFX":     str(ro["data"]),
                "Memo OFX":     ro["memo"],
                "Valor OFX":    ro["valor_ofx"],
                "FITID OFX":    ro.get("fitid",""),
                "_src_idx":     -1,
            })

    return pd.DataFrame(resultado), df_ofx_pix


# ─────────────────────────────────────────────
# FUNÇÕES — GOOGLE SHEETS
# ─────────────────────────────────────────────

def gsheets_url_to_csv(url: str) -> str:
    """
    Converte qualquer URL do Google Sheets para URL de exportação CSV.
    Aceita:
      - URL de edição:    .../spreadsheets/d/ID/edit#gid=123
      - URL de publicação:.../spreadsheets/d/ID/pub?gid=123
      - URL de exportação já formatada
    """
    import re
    url = url.strip()
    # Extrai ID da planilha
    m_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m_id:
        return url  # devolve como está — pode já ser CSV direto
    sheet_id = m_id.group(1)
    # Extrai gid da aba (opcional)
    m_gid = re.search(r"[?&#]gid=(\d+)", url)
    gid_param = f"&gid={m_gid.group(1)}" if m_gid else ""
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv{gid_param}"


def ler_gsheets_como_bytes(url: str) -> "io.BytesIO | None":
    """
    Baixa uma planilha pública do Google Sheets como CSV e devolve
    um objeto BytesIO que imita um arquivo enviado via st.file_uploader.
    Retorna None em caso de erro, exibindo mensagem ao usuário.
    """
    import urllib.request, io as _io
    csv_url = gsheets_url_to_csv(url)
    try:
        with urllib.request.urlopen(csv_url, timeout=15) as resp:
            data = resp.read()
        buf = _io.BytesIO(data)
        buf.name = "gsheets_export.csv"   # parse_pix_* usa .name apenas para ACCTID
        return buf
    except Exception as e:
        st.sidebar.error(f"Erro ao baixar Google Sheets: {e}")
        return None


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.header("📂 Importar Arquivos")
    file_estab = st.file_uploader("📋 Estabelecimentos (.xlsx)", type=["xlsx"])
    file_ofx   = st.file_uploader("Extrato Bancário (.ofx)",     type=["ofx", "OFX"])
    file_rede  = st.file_uploader("Extrato Intermediadora (.xls)", type=["xls", "xlsx", "tsv", "txt"])
    file_caixa   = st.file_uploader("🏪 Relatório de Caixa (.xlsx)", type=["xlsx"])
    # ── PIX POS: arquivo ou Google Sheets ──
    st.markdown("**🔵 PIX POS**")
    _modo_pos = st.radio("Fonte POS:", ["📁 Arquivo", "🌐 Google Sheets"],
                          horizontal=True, key="modo_pos", label_visibility="collapsed")
    if _modo_pos == "📁 Arquivo":
        file_pix_pos = st.file_uploader("PIX POS (.xls)", type=["xls","xlsx"],
                                         label_visibility="collapsed")
    else:
        _url_pos = st.text_input("URL da planilha POS:", key="url_pos",
                                  placeholder="https://docs.google.com/spreadsheets/d/...")
        file_pix_pos = ler_gsheets_como_bytes(_url_pos) if _url_pos.strip() else None

    # ── PIX TEF: arquivo ou Google Sheets ──
    st.markdown("**🟣 PIX TEF**")
    _modo_tef = st.radio("Fonte TEF:", ["📁 Arquivo", "🌐 Google Sheets"],
                          horizontal=True, key="modo_tef", label_visibility="collapsed")
    if _modo_tef == "📁 Arquivo":
        file_pix_tef = st.file_uploader("PIX TEF (.xlsx)", type=["xls","xlsx"],
                                         label_visibility="collapsed")
    else:
        _url_tef = st.text_input("URL da planilha TEF:", key="url_tef",
                                  placeholder="https://docs.google.com/spreadsheets/d/...")
        file_pix_tef = ler_gsheets_como_bytes(_url_tef) if _url_tef.strip() else None

    st.divider()
    st.header("⚙️ Parâmetros")
    tolerancia_dias  = st.slider("Tolerância de data (dias)", 0, 5, 1)
    tolerancia_valor = st.slider("Tolerância de valor (%)",   0, 10, 5) / 100
    col_valor_rede   = st.radio("Comparar OFX com:", ["Valor Líquido", "Valor Bruto"])

    st.divider()
    st.markdown("**📅 Prazo de recebimento estimado**")
    prazo_debito  = st.number_input("Débito — dias úteis após venda", min_value=1, max_value=5,  value=1)
    prazo_credito_modo = st.radio("Crédito — prazo em:", ["Dias úteis", "Dias corridos"], index=1)
    prazo_credito = st.number_input(
        f"Crédito — qtd de {'dias úteis' if prazo_credito_modo == 'Dias úteis' else 'dias corridos'}",
        min_value=1, max_value=60, value=30
    )

    st.divider()
    st.markdown("**Legenda:**")
    st.markdown("✅ Conciliado automaticamente")
    st.markdown("⚠️ Conciliado c/ divergência")
    st.markdown("🔗 Vinculado manualmente")
    st.markdown("❌ Não conciliado")

# ─────────────────────────────────────────────
# ESTADO DA SESSÃO
# ─────────────────────────────────────────────
if "vinculos_manuais" not in st.session_state:
    st.session_state["vinculos_manuais"] = {}
if "vinculos_caixa" not in st.session_state:
    st.session_state["vinculos_caixa"] = []
if "vinculos_pix_pos" not in st.session_state:
    st.session_state["vinculos_pix_pos"] = []
if "vinculos_pix_tef" not in st.session_state:
    st.session_state["vinculos_pix_tef"] = []

# ─────────────────────────────────────────────
# ESTABELECIMENTOS
# ─────────────────────────────────────────────
df_estab = pd.DataFrame()
if file_estab:
    try:
        df_estab = parse_estabelecimentos(file_estab)
    except Exception as e:
        st.sidebar.error(f"Erro ao ler estabelecimentos: {e}")

def lookup_estab(df_estab, key, col):
    """Retorna Fantasia dado um valor na coluna col.
    Normaliza ambos os lados: remove espaços e converte float→int→str
    para garantir match mesmo com '91440580.0' vs '91440580'.
    """
    if df_estab.empty: return ""
    def norm(v):
        s = str(v).strip()
        try:
            return str(int(float(s)))
        except:
            return s
    key_norm = norm(key)
    row = df_estab[df_estab[col].apply(norm) == key_norm]
    return row.iloc[0]["Fantasia"] if not row.empty else ""

# ─────────────────────────────────────────────
# AGUARDA ARQUIVOS
# ─────────────────────────────────────────────
# Arquivos obrigatórios: Lista Estabelecimentos + OFX + Intermediadora
_faltando = []
if not file_estab: _faltando.append("📋 Lista de Estabelecimentos")
if not file_ofx:   _faltando.append("🏦 Extrato Bancário (OFX)")
if not file_rede:  _faltando.append("📊 Extrato Intermediadora (XLS)")

if _faltando:
    st.info("👈 Para iniciar, importe os arquivos obrigatórios na barra lateral:")
    for arq in _faltando:
        st.markdown(f"- **{arq}**")
    with st.expander("ℹ️ Como usar"):
        st.markdown("""
**Arquivos obrigatórios:**
- **Lista de Estabelecimentos (.xlsx)** — Mapeamento ACCTID × Estabelecimento
- **Extrato Bancário (.ofx)** — Exportado pelo banco
- **Extrato Intermediadora (.xls)** — Relatório da adquirente/rede

**Arquivos opcionais (liberam módulos extras):**
- **Relatório de Caixa (.xlsx)** — Libera a aba 🏪 Caixa
- **PIX POS (.xls)** e/ou **PIX TEF (.xlsx)** — Libera a aba 🔵 PIX
        """)
    st.stop()

# ─────────────────────────────────────────────
# PROCESSAMENTO
# ─────────────────────────────────────────────
with st.spinner("Processando arquivos..."):
    try:
        ofx_bytes   = file_ofx.read()
        acctid_ofx  = acctid_do_ofx(ofx_bytes)
        # Lookup feito aqui — df_estab já está carregado neste ponto
        estab_ofx   = lookup_estab(df_estab, acctid_ofx, "ACCTID") if acctid_ofx else ""
        df_ofx_raw  = parse_ofx(ofx_bytes)
        # Se df_estab ainda não foi carregado, tenta novamente após parse
        if not estab_ofx and acctid_ofx and not df_estab.empty:
            estab_ofx = lookup_estab(df_estab, acctid_ofx, "ACCTID")
        if df_ofx_raw.empty:
            st.error("Nenhuma transação encontrada no OFX."); st.stop()
    except Exception as e:
        st.error(f"Erro ao ler OFX: {e}"); st.stop()

    try:
        df_rede_orig = parse_intermediadora_xls(file_rede)
        if df_rede_orig.empty:
            st.error("Nenhuma transação encontrada no arquivo da intermediadora."); st.stop()
        # Identifica estabelecimento:
        # 1. Usa ACCTID do OFX → busca na lista → obtém código ESTABELECIMENTO
        # 2. Filtra df_rede_orig para só esse estabelecimento
        estab_rede  = ""
        cod_estab_filtro = ""
        if not df_estab.empty and acctid_ofx:
            row_e = df_estab[df_estab["ACCTID"].apply(lambda v: str(int(float(str(v)))).strip() if str(v) not in ("","nan") else "") == acctid_ofx]
            if not row_e.empty:
                estab_rede       = row_e.iloc[0]["Fantasia"]
                cod_estab_filtro = str(row_e.iloc[0]["ESTABELECIMENTO"]).strip()
        # Fallback: pega o primeiro da Rede que bater com a lista
        if not estab_rede and "estabelecimento" in df_rede_orig.columns and not df_estab.empty:
            for ev in df_rede_orig["estabelecimento"].dropna().unique():
                found = lookup_estab(df_estab, ev, "ESTABELECIMENTO")
                if found:
                    estab_rede = found
                    cod_estab_filtro = str(ev).strip()
                    break
        # Filtra df_rede_orig pelo estabelecimento identificado
        if cod_estab_filtro and "estabelecimento" in df_rede_orig.columns:
            mask_e = df_rede_orig["estabelecimento"].str.strip() == cod_estab_filtro
            if mask_e.sum() > 0:
                df_rede_orig = df_rede_orig[mask_e].copy()
    except Exception as e:
        st.error(f"Erro ao ler arquivo da intermediadora: {e}"); st.stop()

# Filtros OFX
df_ofx        = df_ofx_raw[~df_ofx_raw["memo"].str.upper().str.contains("SALDO TOTAL", na=False)].copy()
df_ofx_rede   = df_ofx[df_ofx["memo"].str.upper().str.contains("REDE", na=False)].copy()
df_ofx_outros = df_ofx[~df_ofx["memo"].str.upper().str.contains("REDE", na=False)].copy()

if col_valor_rede == "Valor Bruto" and "valor_bruto" in df_rede_orig.columns:
    df_rede_orig["valor_liquido"] = df_rede_orig["valor_bruto"]

df_rede_grupo = agrupar_rede(df_rede_orig)
df_result     = conciliar(df_ofx_rede, df_rede_grupo, tolerancia_dias, tolerancia_valor)

# Lançamentos OFX REDE pendentes:
# Usa fitid direto do df_result (✅/⚠️) + vínculos manuais da sessão
fitids_conciliados = set(
    df_result[df_result["Status"].str.startswith(("✅", "⚠️"))]["fitid_ofx"].dropna()
)
fitids_conciliados.discard("")

fitids_manuais = set(
    info["fitid_ofx"] for info in st.session_state["vinculos_manuais"].values()
    if info.get("fitid_ofx")
)

df_ofx_pendentes = df_ofx_rede[
    ~df_ofx_rede["fitid"].isin(fitids_conciliados | fitids_manuais)
].copy()

# Grupos não conciliados automaticamente
# Apenas grupos com idx_grupo válido (lado Rede não conciliado)
grupos_nao_conc = df_result[
    df_result["Status"].str.contains("❌") &
    df_result["idx_grupo"].notna()
].copy()
grupos_nao_conc["idx_grupo"] = grupos_nao_conc["idx_grupo"].astype(int)
grupos_vinculados_manual = set(st.session_state["vinculos_manuais"].keys())
# Índices de transações já vinculadas manualmente (virtuais)
transacoes_ja_vinculadas = set()
for info in st.session_state["vinculos_manuais"].values():
    for idx_t in info.get("idx_transacoes", []):
        transacoes_ja_vinculadas.add(idx_t)

# ─────────────────────────────────────────────
# KPIs
# ─────────────────────────────────────────────
total_grupos    = len(df_rede_grupo)
total_conc_auto = df_result["Status"].str.contains("✅|⚠️", regex=True).sum()
total_manual    = len(st.session_state["vinculos_manuais"])
total_pendentes = max(len(grupos_nao_conc) - total_manual, 0)
pct = lambda n: f"{n/total_grupos*100:.1f}%" if total_grupos else "0%"

c1, c2, c3, c4 = st.columns(4)
c1.metric("📋 Grupos Intermediadora",  total_grupos)
c2.metric("✅ Conciliados auto",        f"{total_conc_auto} ({pct(total_conc_auto)})")
c3.metric("🔗 Vinculados manualmente", f"{total_manual} ({pct(total_manual)})")
c4.metric("❌ Pendentes",              f"{total_pendentes} ({pct(total_pendentes)})")

val_ofx_rede  = df_ofx_rede["valor_ofx"].apply(lambda v: v if v > 0 else 0).sum()
val_rede_bruto = df_rede_orig["valor_bruto"].sum()  if "valor_bruto"   in df_rede_orig.columns else 0
val_rede_liq   = df_rede_orig["valor_liquido"].sum() if "valor_liquido" in df_rede_orig.columns else 0

cv1, cv2, cv3, cv4 = st.columns(4)
cv1.metric("💰 OFX — Lançamentos REDE",          f"R$ {val_ofx_rede:,.2f}")
cv2.metric("💳 Intermediadora — Valor Bruto",     f"R$ {val_rede_bruto:,.2f}")
cv3.metric("🏦 Intermediadora — Valor Líquido",   f"R$ {val_rede_liq:,.2f}")
cv4.metric("📊 Diferença OFX × Líquido",
           f"R$ {val_ofx_rede - val_rede_liq:,.2f}",
           delta=f"{((val_ofx_rede - val_rede_liq)/val_rede_liq*100):.2f}%" if val_rede_liq else None)

st.divider()

# Banner de estabelecimento
if estab_ofx and estab_rede:
    match = "✅" if estab_ofx == estab_rede else "⚠️"
    label = estab_ofx
    _estab_placeholder.info(f"🏪 Estabelecimento: **{label}** {match}  |  ACCTID OFX: `{acctid_ofx}`")
elif estab_ofx:
    _estab_placeholder.info(f"🏪 Estabelecimento: **{estab_ofx}** (via OFX)  |  ACCTID OFX: `{acctid_ofx}`")
elif estab_rede:
    if acctid_ofx:
        # ACCTID lido mas não encontrado na lista — provavelmente lista não importada ainda
        _estab_placeholder.warning(
            f"⚠️ ACCTID `{acctid_ofx}` não encontrado na lista de estabelecimentos.  "
            f"Intermediadora identificou: **{estab_rede}** — importe a Lista de Estabelecimentos para cruzar."
        )
    else:
        _estab_placeholder.info(f"🏪 Estabelecimento: **{estab_rede}** (via intermediadora)")
elif acctid_ofx:
    _estab_placeholder.warning(
        f"⚠️ ACCTID `{acctid_ofx}` encontrado no OFX mas lista de estabelecimentos não importada."
    )

# ─────────────────────────────────────────────
# ABAS PRINCIPAIS
# ─────────────────────────────────────────────
# Monta lista de abas dinamicamente conforme arquivos disponíveis
_nomes_abas = [
    "🏪 Caixa",
    "🔍 Conciliação",
    "📋 Detalhe por Transação",
    "🔗 Vinculação Manual",
    "📄 Outros Lançamentos OFX",
]
_tem_pix = file_pix_pos or file_pix_tef
if _tem_pix:
    _nomes_abas.append("🔵 PIX")

_abas = st.tabs(_nomes_abas)

# Mapeia nomes para variáveis — Caixa só ativo se arquivo incluído
aba_caixa   = _abas[0]
aba_result  = _abas[1]
aba_detalhe = _abas[2]
aba_manual  = _abas[3]
aba_outros  = _abas[4]
aba_pix_pos = _abas[5] if _tem_pix else None

# ════════════════════════════════════════════
# ABA 1 — RESULTADO DA CONCILIAÇÃO
# ════════════════════════════════════════════
with aba_result:
    st.subheader("Resultado — Grupos (Intermediadora × OFX)")

    # Aplica vínculos manuais na exibição
    # Vínculos virtuais (por idx_transacao) não têm idx_grupo — são ignorados aqui,
    # pois o status deles aparece no Detalhe por Transação
    df_result_display = df_result.copy()
    for ig, info in st.session_state["vinculos_manuais"].items():
        if info.get("virtual"):
            continue  # vínculo por transação individual — sem idx_grupo no resultado
        try:
            ig_int = int(ig)
        except (ValueError, TypeError):
            continue
        mask = df_result_display["idx_grupo"] == ig_int
        df_result_display.loc[mask, "Status"]    = info["status"]
        df_result_display.loc[mask, "Memo OFX"]  = info.get("memo_ofx", "")
        df_result_display.loc[mask, "Valor OFX"] = info.get("valor_ofx", "")
        df_result_display.loc[mask, "Data OFX"]  = info.get("data_ofx", "")

    status_opcoes = df_result_display["Status"].unique().tolist()
    filtro = st.multiselect("Filtrar por status:", options=status_opcoes,
                             default=status_opcoes, key="filtro_resultado")
    df_filtrado = df_result_display[df_result_display["Status"].isin(filtro)]

    def fmt_brl(v):
        try:
            if v == "" or pd.isna(v): return ""
            return f"R$ {float(v):,.2f}"
        except: return v

    df_show = df_filtrado.drop(columns=["idx_grupo", "fitid_ofx"], errors="ignore").copy()
    for c in ["Valor OFX", "Valor Bruto Rede", "Valor Líq. Rede", "Diferença (R$)"]:
        if c in df_show.columns:
            df_show[c] = df_show[c].apply(fmt_brl)

    st.dataframe(df_show, use_container_width=True, height=420)

# ════════════════════════════════════════════
# ABA 2 — DETALHE POR TRANSAÇÃO
# ════════════════════════════════════════════
with aba_detalhe:
    st.subheader("Detalhe por Transação Individual")
    st.caption("Cada transação herda o status do seu grupo (Data + Bandeira + Tipo).")

    df_detalhe = build_status_transacao(
        df_rede_orig, df_rede_grupo, df_result,
        st.session_state["vinculos_manuais"]
    )

    status_det_opcoes = df_detalhe["Status Conciliação"].unique().tolist()
    filtro_det = st.multiselect("Filtrar por status:", options=status_det_opcoes,
                                 default=status_det_opcoes, key="filtro_detalhe")
    df_det_filtrado = df_detalhe[df_detalhe["Status Conciliação"].isin(filtro_det)]

    cols_show = ["Status Conciliação", "data", "bandeira", "tipo_norm", "tipo",
                 "valor_bruto", "taxa_final", "taxa_pct", "valor_liquido",
                 "cv", "estabelecimento", "status_adq",
                 "Memo OFX Vinculado", "Valor OFX Vinculado"]
    cols_show = [c for c in cols_show if c in df_det_filtrado.columns]

    rename_det = {
        "Status Conciliação":  "Status",
        "data":                "Data",
        "bandeira":            "Bandeira",
        "tipo_norm":           "Tipo",
        "tipo":                "Descrição",
        "valor_bruto":         "Valor Bruto",
        "taxa_final":          "Taxa (R$)",
        "taxa_pct":            "Taxa (%)",
        "valor_liquido":       "Valor Líquido",
        "cv":                  "C.V.",
        "estabelecimento":     "Estabelecimento",
        "status_adq":          "Status Adquirente",
        "Memo OFX Vinculado":  "Memo OFX",
        "Valor OFX Vinculado": "Valor OFX",
    }

    df_det_show = df_det_filtrado[cols_show].rename(columns=rename_det).copy()
    for c in ["Valor Bruto", "Valor Líquido", "Taxa (R$)"]:
        if c in df_det_show.columns:
            df_det_show[c] = df_det_show[c].apply(
                lambda v: f"R$ {v:,.2f}" if pd.notna(v) and v != "" else ""
            )

    st.dataframe(df_det_show, use_container_width=True, height=450)

    with st.expander("📊 Resumo por status"):
        resumo_status = df_detalhe["Status Conciliação"].value_counts().reset_index()
        resumo_status.columns = ["Status", "Qtd Transações"]
        st.dataframe(resumo_status, use_container_width=True)

# ════════════════════════════════════════════
# ABA 3 — VINCULAÇÃO MANUAL
# ════════════════════════════════════════════
with aba_manual:
    st.subheader("🔗 Vinculação Manual de Lançamentos")
    st.info(
        "Selecione um **lançamento OFX pendente**, depois marque as **transações individuais** "
        "da intermediadora que compõem esse valor. Somente lançamentos ❌ estão disponíveis.",
        icon="ℹ️"
    )

    # Índices de transações já vinculadas manualmente (não podem ser reusadas)
    transacoes_vinculadas = set()
    for info in st.session_state["vinculos_manuais"].values():
        for idx_t in info.get("idx_transacoes", []):
            transacoes_vinculadas.add(idx_t)

    if df_ofx_pendentes.empty:
        st.success("✅ Não há lançamentos OFX pendentes para vinculação manual.")
    else:
        # ── 1. Seleção do lançamento OFX ────────
        st.markdown("#### 1. Selecione o lançamento OFX pendente")

        mapa_ofx = {}
        for _, row in df_ofx_pendentes.iterrows():
            label = f"{row['data']}  |  {row['memo']}  |  R$ {abs(row['valor_ofx']):,.2f}"
            mapa_ofx[label] = row

        sel_ofx_label = st.selectbox("Lançamento OFX:", list(mapa_ofx.keys()), key="sel_ofx")
        sel_ofx_row   = mapa_ofx.get(sel_ofx_label)

        if sel_ofx_row is None:
            st.warning("Selecione um lançamento OFX para continuar.")
            st.stop()

        val_ofx_s = float(sel_ofx_row["valor_ofx"])

        # ── Detecta bandeira e tipo sugeridos pelo memo do OFX ──
        bandeira_ofx, tipo_ofx = detectar_bandeira_tipo(sel_ofx_row["memo"])

        # ── 2. Selecione as transações da intermediadora ──
        st.markdown("#### 2. Selecione as transações da intermediadora")

        # Transações disponíveis = não conciliadas auto nem vinculadas manual
        _grupos_auto_ok = set(
            df_result[df_result["Status"].str.startswith(("✅", "⚠️"))]["idx_grupo"]
            .dropna().astype(int)
        )
        idxs_bloqueados = set()
        for ig in _grupos_auto_ok:
            row_g = df_rede_grupo[df_rede_grupo["idx_grupo"] == ig]
            if not row_g.empty:
                idxs_bloqueados.update(row_g.iloc[0]["idx_transacao"])
        for info in st.session_state["vinculos_manuais"].values():
            idxs_bloqueados.update(info.get("idx_transacoes", []))

        df_trans_all = df_rede_orig[
            ~df_rede_orig["idx_transacao"].isin(idxs_bloqueados)
        ].copy().reset_index(drop=True)

        if df_trans_all.empty:
            st.warning("Não há transações disponíveis para vinculação.")
            st.stop()

        # ── Filtros ──────────────────────────────────────────
        # Pré-seleciona bandeira/tipo detectados no memo do OFX
        bandeiras_disp = sorted(df_trans_all["bandeira"].dropna().unique().tolist())
        tipos_disp     = sorted(df_trans_all["tipo_norm"].dropna().unique().tolist())
        datas_disp     = sorted(df_trans_all["data"].dropna().unique().tolist())

        # Sugestão automática baseada no memo OFX
        sugestao_bandeira = [b for b in bandeiras_disp if bandeira_ofx != "OUTROS" and b == bandeira_ofx]
        sugestao_tipo     = [t for t in tipos_disp     if tipo_ofx     != "OUTROS" and t == tipo_ofx]

        with st.expander("🔎 Filtros" + (f" — sugestão automática: **{bandeira_ofx} {tipo_ofx}**" if sugestao_bandeira else ""), expanded=True):
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                fil_bandeira = st.multiselect(
                    "Bandeira:", bandeiras_disp,
                    default=sugestao_bandeira if sugestao_bandeira else bandeiras_disp,
                    key="fil_bandeira"
                )
            with fc2:
                fil_tipo = st.multiselect(
                    "Tipo:", tipos_disp,
                    default=sugestao_tipo if sugestao_tipo else tipos_disp,
                    key="fil_tipo"
                )
            with fc3:
                data_min = min(datas_disp)
                data_max = max(datas_disp)
                fil_datas = st.date_input(
                    "Período:",
                    value=(data_min, data_max),
                    min_value=data_min,
                    max_value=data_max,
                    key="fil_datas",
                    format="DD/MM/YYYY",
                )

        # Aplica filtros
        df_trans_disponiveis = df_trans_all[
            df_trans_all["bandeira"].isin(fil_bandeira) &
            df_trans_all["tipo_norm"].isin(fil_tipo)
        ].copy()

        if isinstance(fil_datas, (list, tuple)) and len(fil_datas) == 2:
            d_ini, d_fim = fil_datas
            df_trans_disponiveis = df_trans_disponiveis[
                (df_trans_disponiveis["data"] >= d_ini) &
                (df_trans_disponiveis["data"] <= d_fim)
            ]

        df_trans_disponiveis = df_trans_disponiveis.reset_index(drop=True)

        st.caption(f"**{len(df_trans_disponiveis)}** transação(ões) exibida(s) após filtros "
                   f"(total disponível: {len(df_trans_all)})")

        if df_trans_disponiveis.empty:
            st.info("Nenhuma transação encontrada com os filtros aplicados.")
        else:
            # ── Tabela com checkbox (data_editor) ──
            cols_tabela = ["data", "bandeira", "tipo_norm", "cv",
                           "valor_bruto", "taxa_final", "valor_liquido"]
            cols_tabela = [c for c in cols_tabela if c in df_trans_disponiveis.columns]

            df_editor = df_trans_disponiveis[cols_tabela].copy()

            # Coluna de previsão de recebimento
            df_editor["Prev. Recebimento"] = df_editor.apply(
                lambda r: calcular_previsao(
                    r["data"], r["tipo_norm"],
                    prazo_debito, prazo_credito, prazo_credito_modo
                ), axis=1
            )

            df_editor["data"] = df_editor["data"].apply(
                lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) and hasattr(d, "strftime") else str(d)
            )

            # ── Controle de "selecionar todos" via session_state ──
            chave_todos = f"sel_todos_{sel_ofx_label}"
            if chave_todos not in st.session_state:
                st.session_state[chave_todos] = False

            # Botões acima da tabela
            btn_col1, btn_col2, _ = st.columns([1, 1, 4])
            with btn_col1:
                if st.button("☑️ Selecionar todos", key="btn_sel_todos", use_container_width=True):
                    st.session_state[chave_todos] = True
            with btn_col2:
                if st.button("🔲 Desmarcar todos", key="btn_des_todos", use_container_width=True):
                    st.session_state[chave_todos] = False

            # Define valor inicial da coluna checkbox
            valor_inicial = st.session_state[chave_todos]
            df_editor.insert(0, "✔", valor_inicial)

            # Formata valores para exibição
            for c, col in [("valor_bruto", "Valor Bruto"), ("taxa_final", "Taxa (R$)"),
                           ("valor_liquido", "Valor Líquido")]:
                if c in df_editor.columns:
                    df_editor[col] = df_editor[c].apply(lambda v: f"R$ {v:,.2f}")
                    df_editor = df_editor.drop(columns=[c])

            df_editor = df_editor.rename(columns={
                "data": "Data Venda", "bandeira": "Bandeira",
                "tipo_norm": "Tipo", "cv": "C.V.",
            })

            edited = st.data_editor(
                df_editor,
                column_config={
                    "✔":                  st.column_config.CheckboxColumn("✔", help="Marque para incluir", width="small"),
                    "Data Venda":         st.column_config.TextColumn("Data Venda",    width="small"),
                    "Prev. Recebimento":  st.column_config.TextColumn("Prev. Recebimento", width="small",
                                          help="Estimativa de quando o valor cai no banco"),
                    "Bandeira":           st.column_config.TextColumn("Bandeira",      width="small"),
                    "Tipo":               st.column_config.TextColumn("Tipo",          width="small"),
                    "C.V.":               st.column_config.TextColumn("C.V.",          width="medium"),
                    "Valor Bruto":        st.column_config.TextColumn("Valor Bruto",   width="medium"),
                    "Taxa (R$)":          st.column_config.TextColumn("Taxa (R$)",     width="small"),
                    "Valor Líquido":      st.column_config.TextColumn("Valor Líquido", width="medium"),
                },
                disabled=["Data Venda", "Prev. Recebimento", "Bandeira", "Tipo", "C.V.",
                          "Valor Bruto", "Taxa (R$)", "Valor Líquido"],
                hide_index=True,
                use_container_width=True,
                key="editor_transacoes",
                height=min(500, 45 + len(df_editor) * 35),
            )

            # Recupera índices das linhas marcadas
            idx_marcados = edited[edited["✔"]].index.tolist()
            df_sel = df_trans_disponiveis.iloc[idx_marcados]
            total_liq_sel   = df_sel["valor_liquido"].sum()
            total_bruto_sel = df_sel["valor_bruto"].sum()
            total_taxa_sel  = df_sel["taxa_final"].sum()
            selecionadas    = idx_marcados   # usado apenas para len()

            # ── 3. Conferência ───────────────────
            st.markdown("#### 3. Conferência de valores")
            diff_val = abs(val_ofx_s) - total_liq_sel
            diff_pct = (diff_val / abs(val_ofx_s) * 100) if val_ofx_s else 0

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Valor OFX",         f"R$ {abs(val_ofx_s):,.2f}")
            m2.metric("Selecionado Bruto",  f"R$ {total_bruto_sel:,.2f}")
            m3.metric("Selecionado Líq.",   f"R$ {total_liq_sel:,.2f}",
                      help="Deve ser igual ou próximo ao Valor OFX")
            m4.metric("Diferença",          f"R$ {diff_val:,.2f}",
                      delta=f"{diff_pct:.2f}%" if selecionadas else None,
                      delta_color="off" if abs(diff_val) < 0.01 else "inverse")

            obs = st.text_input("Observação (opcional):", key="obs_manual",
                                placeholder="Ex: transações do dia 04/01 referentes ao OFX de 06/01")

            # Validações
            nenhuma_sel   = len(selecionadas) == 0
            diff_bloqueio = abs(diff_pct) > 10 and not nenhuma_sel

            if nenhuma_sel:
                st.info("Selecione ao menos uma transação para confirmar o vínculo.")
            elif diff_bloqueio:
                st.error(f"⛔ Diferença de {diff_pct:.1f}% acima de 10%. "
                          "Revise as transações selecionadas.")
            elif abs(diff_val) > 0.01:
                st.warning(f"⚠️ Diferença de R$ {diff_val:,.2f} ({diff_pct:.2f}%). "
                            "O vínculo será marcado como divergente.")

            col_btn, _ = st.columns([1, 3])
            with col_btn:
                confirmar = st.button(
                    "✅ Confirmar Vínculo", type="primary",
                    use_container_width=True,
                    disabled=(nenhuma_sel or diff_bloqueio)
                )

            if confirmar:
                # Cria um grupo virtual com as transações selecionadas
                idx_virtual = f"manual_{len(st.session_state['vinculos_manuais'])}"
                status_manual = "🔗 Vinculado Manualmente" if abs(diff_val) < 0.01 else "🔗 Vinculado c/ Divergência"

                st.session_state["vinculos_manuais"][idx_virtual] = {
                    "status":         status_manual,
                    "memo_ofx":       sel_ofx_row["memo"],
                    "fitid_ofx":      sel_ofx_row.get("fitid", ""),   # chave única do OFX
                    "valor_ofx":      sel_ofx_row["valor_ofx"],
                    "data_ofx":       sel_ofx_row["data"],
                    "observacao":     obs,
                    "diff_valor":     diff_val,
                    "idx_transacoes": list(df_sel["idx_transacao"]),
                    "total_liq":      total_liq_sel,
                    "total_bruto":    total_bruto_sel,
                    "total_taxa":     total_taxa_sel,
                    "virtual":        True,
                }
                st.success(f"✅ {len(selecionadas)} transação(ões) vinculada(s) ao OFX **{sel_ofx_row['memo']}**!")
                st.rerun()

    # ── Lista de vínculos registrados ───────────
    if st.session_state["vinculos_manuais"]:
        st.divider()
        st.markdown("#### Vínculos manuais registrados nesta sessão")

        registros = []
        for chave, info in st.session_state["vinculos_manuais"].items():
            qtd_trans = len(info.get("idx_transacoes", []))
            registros.append({
                "Status":          info["status"],
                "Memo OFX":        info.get("memo_ofx", ""),
                "Data OFX":        str(info.get("data_ofx", "")),
                "Valor OFX":       f"R$ {abs(float(info.get('valor_ofx', 0))):,.2f}",
                "Valor Líq. Sel.": f"R$ {float(info.get('total_liq', 0)):,.2f}",
                "Diferença":       f"R$ {float(info.get('diff_valor', 0)):,.2f}",
                "Qtd Transações":  qtd_trans,
                "Observação":      info.get("observacao", ""),
            })

        df_vinculos = pd.DataFrame(registros)
        st.dataframe(df_vinculos, use_container_width=True)

        col_limpar, _ = st.columns([1, 3])
        with col_limpar:
            if st.button("🗑️ Limpar todos os vínculos manuais", type="secondary",
                         use_container_width=True):
                st.session_state["vinculos_manuais"] = {}
                st.rerun()

# ════════════════════════════════════════════
# ABA 4 — OUTROS LANÇAMENTOS OFX
# ════════════════════════════════════════════
with aba_outros:
    st.subheader("Outros Lançamentos OFX — Não Relacionados à Rede")
    if df_ofx_outros.empty:
        st.info("Nenhum lançamento OFX fora do escopo da Rede.")
    else:
        df_out = df_ofx_outros[["data", "valor_ofx", "memo"]].copy()
        df_out["valor_ofx"] = df_out["valor_ofx"].apply(lambda v: f"R$ {v:,.2f}")
        df_out = df_out.rename(columns={"data": "Data", "valor_ofx": "Valor", "memo": "Memo"})
        st.dataframe(df_out, use_container_width=True)
        total_outros = df_ofx_outros["valor_ofx"].apply(lambda v: v if v > 0 else 0).sum()
        st.metric("Total créditos", f"R$ {total_outros:,.2f}")

st.divider()

# ════════════════════════════════════════════
# ABA 5 — CAIXA
# ════════════════════════════════════════════
with aba_caixa:
    st.subheader("🏪 Conciliação Caixa × Intermediadora")

    if not file_caixa:
        st.info("👈 Importe o Relatório de Caixa (.xlsx) na barra lateral para usar este módulo.")
    else:
        # Parse do caixa
        try:
            nome_arquivo_caixa = file_caixa.name
            filial_caixa = nome_filial_do_arquivo(nome_arquivo_caixa)
            df_caixa = parse_caixa(file_caixa)
        except Exception as e:
            st.error(f"Erro ao ler arquivo de caixa: {e}")
            st.stop()

        # Identifica estabelecimento do caixa pelo nome do arquivo vs lista
        estab_caixa = ""
        if not df_estab.empty:
            for _, er in df_estab.iterrows():
                if str(er["Fantasia"]).upper() == filial_caixa:
                    estab_caixa = er["Fantasia"]
                    break
            if not estab_caixa:
                estab_caixa = filial_caixa

        st.info(f"🏪 Filial: **{estab_caixa}**  |  Arquivo: `{nome_arquivo_caixa}`")

        # ── KPIs do caixa ──────────────────────
        total_cx       = len(df_caixa)
        total_cartoes  = df_caixa[df_caixa["forma_norm"].isin(["CREDITO","DEBITO"])]
        total_pix_cx   = df_caixa[df_caixa["forma_norm"] == "PIX"]
        outros_cx      = df_caixa[~df_caixa["forma_norm"].isin(["CREDITO","DEBITO","PIX"])]

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("📋 Total lançamentos", total_cx)
        k2.metric("💳 Cartões",  f"{len(total_cartoes)} — R$ {total_cartoes['valor'].sum():,.2f}")
        k3.metric("🔵 PIX",      f"{len(total_pix_cx)} — R$ {total_pix_cx['valor'].sum():,.2f}")
        k4.metric("📦 Outros",   f"{len(outros_cx)} — R$ {outros_cx['valor'].sum():,.2f}")

        st.divider()

        # ── Conciliação Caixa × Rede (executa primeiro para enriquecer o resumo) ──
        df_rede_para_caixa = df_rede_orig.copy()
        if "estabelecimento" in df_rede_para_caixa.columns and not df_estab.empty and estab_caixa:
            row_estab = df_estab[df_estab["Fantasia"].str.upper() == estab_caixa.upper()]
            if not row_estab.empty:
                cod_estab = str(row_estab.iloc[0]["ESTABELECIMENTO"])
                mask_estab = df_rede_para_caixa["estabelecimento"].astype(str).str.strip() == cod_estab
                df_rede_para_caixa = df_rede_para_caixa[mask_estab]

        with st.spinner("Conciliando caixa com intermediadora..."):
            df_conc_caixa = conciliar_caixa_rede(df_caixa, df_rede_para_caixa)

        # Adiciona índice permanente para referência nos vínculos manuais
        df_conc_caixa = df_conc_caixa.reset_index(drop=True)
        df_conc_caixa["_idx"] = df_conc_caixa.index

        # Aplica vínculos manuais da sessão
        for vinc in st.session_state["vinculos_caixa"]:
            for i in vinc.get("idx_caixa", []) + vinc.get("idx_rede", []):
                if i < len(df_conc_caixa):
                    df_conc_caixa.at[i, "Status"] = "🔗 Vinculado Manualmente"

        # ── Resumo por caixa ───────────────────
        st.markdown("#### Resumo por Caixa e Forma de Pagamento")

        # ── Helpers de status de conciliação ──
        def status_conc_por_caixa(df_conc: pd.DataFrame, forma: str) -> dict:
            """Retorna dict {caixa: emoji} — prioridade: ❌ > ⚠️ > 🔗 > ✅."""
            resultado = {}
            sub = df_conc[df_conc["forma_norm"] == forma]
            for caixa, grp in sub.groupby("Caixa"):
                ss = grp["Status"].tolist()
                if any("❌" in s for s in ss):   resultado[caixa] = "❌"
                elif any("⚠️" in s for s in ss): resultado[caixa] = "⚠️"
                elif any("🔗" in s for s in ss): resultado[caixa] = "🔗"
                else:                             resultado[caixa] = "✅"
            return resultado

        def status_geral_caixa(cx, status_dict_list):
            """Combina múltiplos dicts de status para um caixa → emoji mais grave."""
            emojis = [d.get(cx, "") for d in status_dict_list if cx in d]
            if not emojis: return "—"
            if any("❌" in e for e in emojis):   return "❌"
            if any("⚠️" in e for e in emojis):  return "⚠️"
            if any("🔗" in e for e in emojis):   return "🔗"
            return "✅"

        status_cred = status_conc_por_caixa(df_conc_caixa, "CREDITO")
        status_deb  = status_conc_por_caixa(df_conc_caixa, "DEBITO")
        # PIX: usa status geral armazenado pela aba PIX POS (se já foi processada)
        _spg = st.session_state.get("_status_pix_geral", "")
        # Aplica o mesmo status a todos os caixas que têm PIX (PIX POS não tem nº caixa)
        _caixas_com_pix = df_caixa[df_caixa["forma_norm"]=="PIX"]["Caixa"].unique().tolist()
        status_pix = {cx: _spg for cx in _caixas_com_pix} if _spg else {}

        # ── Pivot numérico base ──
        pivot_num = df_caixa.groupby(["Caixa", "forma_norm"])["valor"].sum().unstack(fill_value=0)
        pivot_num.columns.name = None
        pivot_num = pivot_num.reset_index()

        # Normaliza nomes das formas para lookup (uppercase)
        forma_map = {c: c for c in pivot_num.columns if c != "Caixa"}

        def val(row, forma):
            """Retorna valor numérico de uma forma ou 0."""
            return float(row.get(forma, 0) or 0)

        # ── Monta DataFrame final com colunas na ordem solicitada ──
        rows_disp = []
        conc_hint = "✅ Tudo conciliado  |  ⚠️ Divergência  |  🔗 Vinculado manualmente  |  ❌ Não conciliado  |  — Sem transações"

        for _, row in pivot_num.iterrows():
            cx = row["Caixa"]

            v_din    = val(row, "DINHEIRO")
            v_cred   = val(row, "CREDITO")
            v_deb    = val(row, "DEBITO")
            v_pix    = val(row, "PIX")
            v_book   = val(row, "BOOKING")
            v_ggr    = val(row, "GUIA GO RECEBER")
            v_ggt    = val(row, "GUIA GO TAXA")
            v_dup    = val(row, "DUPLICATA")
            v_cort   = val(row, "CORTESIA")

            v_receita = v_din + v_cred + v_deb + v_pix + v_book + v_ggr
            v_total   = v_receita + v_ggt + v_dup + v_cort

            sc = status_cred.get(cx, "—")
            sd = status_deb.get(cx, "—")
            sp = status_pix.get(cx, "—")
            sg = status_geral_caixa(cx, [status_cred, status_deb, status_pix])

            def fmt(v): return f"R$ {v:,.2f}" if v else "R$ 0,00"

            rows_disp.append({
                "Nº Caixa":        cx,
                "Conc.":           sg,
                "Dinheiro":        fmt(v_din),
                "Crédito":         fmt(v_cred),
                "Conc. Créd.":     sc,
                "Débito":          fmt(v_deb),
                "Conc. Déb.":      sd,
                "PIX":             fmt(v_pix),
                "Conc. PIX":       sp,
                "Booking":         fmt(v_book),
                "Guia Go Rec.":    fmt(v_ggr),
                "Valor Receita":   fmt(v_receita),
                "Guia Go Taxa":    fmt(v_ggt),
                "Duplicata":       fmt(v_dup),
                "Cortesia":        fmt(v_cort),
                "Valor Total":     fmt(v_total),
            })

        df_pivot_disp = pd.DataFrame(rows_disp)

        # ── column_config: oculta nada, TextColumn em tudo, status pequenos ──
        COLS_STATUS  = {"Conc.", "Conc. Créd.", "Conc. Déb.", "Conc. PIX"}
        COLS_VALOR   = {"Nº Caixa", "Dinheiro", "Crédito", "Débito", "PIX",
                        "Booking", "Guia Go Rec.", "Valor Receita",
                        "Guia Go Taxa", "Duplicata", "Cortesia", "Valor Total"}
        col_cfg = {}
        for c in df_pivot_disp.columns:
            if c in COLS_STATUS:
                col_cfg[c] = st.column_config.TextColumn(c, width="small",
                    help=conc_hint)
            else:
                col_cfg[c] = st.column_config.TextColumn(c, width="medium")

        st.dataframe(df_pivot_disp, use_container_width=True,
                     column_config=col_cfg, hide_index=True)

        st.caption(conc_hint)

        # Totais gerais
        tc1, tc2, tc3 = st.columns(3)
        tc1.metric("Total Crédito",
                   f"R$ {df_caixa[df_caixa['forma_norm']=='CREDITO']['valor'].sum():,.2f}")
        tc2.metric("Total Débito",
                   f"R$ {df_caixa[df_caixa['forma_norm']=='DEBITO']['valor'].sum():,.2f}")
        tc3.metric("Total Geral",
                   f"R$ {df_caixa['valor'].sum():,.2f}")

        st.divider()

        # ── Conciliação Cartões — detalhes ─────
        st.markdown("#### Conciliação Cartões — Caixa × Intermediadora")

        # KPIs da conciliação
        n_ok   = df_conc_caixa["Status"].str.startswith("✅").sum()
        n_div  = df_conc_caixa["Status"].str.startswith("⚠️").sum()
        n_err  = df_conc_caixa["Status"].str.startswith("❌").sum()
        tot_cc = len(df_conc_caixa)

        p1, p2, p3, p4 = st.columns(4)
        p1.metric("📋 Total comparado",  tot_cc)
        p2.metric("✅ Conciliados",       f"{n_ok} ({n_ok/tot_cc*100:.1f}%)" if tot_cc else "0")
        p3.metric("⚠️ Divergentes",      f"{n_div} ({n_div/tot_cc*100:.1f}%)" if tot_cc else "0")
        p4.metric("❌ Não encontrados",   f"{n_err} ({n_err/tot_cc*100:.1f}%)" if tot_cc else "0")

        # Filtro de status
        status_cx = df_conc_caixa["Status"].unique().tolist()
        filtro_cx = st.multiselect("Filtrar status:", status_cx, default=status_cx, key="filtro_cx")
        df_cx_show = df_conc_caixa[df_conc_caixa["Status"].isin(filtro_cx)].copy()

        # Formata valores para exibição
        df_cx_disp = df_cx_show.drop(columns=["_idx","forma_norm"], errors="ignore").copy()
        for c in ["Valor Caixa", "Valor Rede"]:
            df_cx_disp[c] = df_cx_disp[c].apply(
                lambda v: f"R$ {float(v):,.2f}" if str(v) not in ("", "nan") else ""
            )

        st.dataframe(df_cx_disp, use_container_width=True, height=400)

        # ════════════════════════════════════════
        # VINCULAÇÃO MANUAL — CAIXA
        # ════════════════════════════════════════
        st.divider()
        st.markdown("#### 🔗 Vinculação Manual")

        # Separa pendentes (não-✅ e não-🔗)
        mask_pend = ~df_conc_caixa["Status"].str.startswith("✅")
        mask_pend &= ~df_conc_caixa["Status"].str.startswith("🔗")
        df_pend = df_conc_caixa[mask_pend].copy()

        if df_pend.empty:
            st.success("Todos os registros estão conciliados ou vinculados.")
        else:
            n_pend = len(df_pend)
            st.info(f"{n_pend} registro(s) pendente(s) de conciliação manual.")

            # Separa lado Caixa e lado Rede
            df_lado_cx   = df_pend[df_pend["Status"] != "❌ Não encontrado no Caixa"].copy()
            df_lado_rede = df_pend[df_pend["Status"] == "❌ Não encontrado no Caixa"].copy()

            col_esq, col_dir = st.columns(2)

            # ── Lado Caixa ──────────────────────
            with col_esq:
                st.markdown("**1. Selecione do lado Caixa:**")
                if df_lado_cx.empty:
                    st.info("Nenhum pendente do lado Caixa.")
                    sel_idx_cx = []
                else:
                    df_ed_cx = df_lado_cx[["_idx","Status","Caixa","Forma","Data/Hora","Valor Caixa","AutExtRef"]].copy()
                    df_ed_cx = df_ed_cx.reset_index(drop=True)  # índice 0..N para alinhar com editor
                    df_ed_cx["Valor Caixa"] = df_ed_cx["Valor Caixa"].apply(
                        lambda v: f"R$ {float(v):,.2f}" if str(v) not in ("","nan") else "")
                    df_ed_cx.insert(0, "✔", False)
                    edited_cx = st.data_editor(
                        df_ed_cx,
                        column_config={
                            "✔":          st.column_config.CheckboxColumn("✔",        width="small"),
                            "_idx":       None,   # oculta coluna técnica
                            "Status":     st.column_config.TextColumn("Status",       width="medium"),
                            "Caixa":      st.column_config.TextColumn("Caixa",        width="small"),
                            "Forma":      st.column_config.TextColumn("Forma",        width="small"),
                            "Data/Hora":  st.column_config.TextColumn("Data/Hora",    width="medium"),
                            "Valor Caixa":st.column_config.TextColumn("Valor",        width="medium"),
                            "AutExtRef":  st.column_config.TextColumn("AutExtRef",    width="medium"),
                        },
                        disabled=["_idx","Status","Caixa","Forma","Data/Hora","Valor Caixa","AutExtRef"],
                        hide_index=True, use_container_width=True,
                        key="ed_vinc_cx",
                        height=min(400, 45 + len(df_ed_cx)*35),
                    )
                    # Lê _idx direto do editor — imune a dessincronias de índice
                    sel_idx_cx = edited_cx[edited_cx["✔"]]["_idx"].astype(int).tolist()
                    if sel_idx_cx:
                        total_sel_cx = df_conc_caixa[df_conc_caixa["_idx"].isin(sel_idx_cx)]["Valor Caixa"].apply(
                            lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                        st.metric("Selecionado Caixa", f"R$ {total_sel_cx:,.2f}",
                                  help=f"{len(sel_idx_cx)} registro(s)")

            # ── Lado Rede ───────────────────────
            with col_dir:
                st.markdown("**2. Selecione do lado Rede:**")
                if df_lado_rede.empty:
                    st.info("Nenhum pendente do lado Rede.")
                    sel_idx_rede = []
                else:
                    df_ed_rede = df_lado_rede[["_idx","Status","C.V. Rede","Data Rede","Bandeira Rede","Valor Rede"]].copy()
                    df_ed_rede = df_ed_rede.reset_index(drop=True)  # índice 0..N para alinhar com editor
                    df_ed_rede["Valor Rede"] = df_ed_rede["Valor Rede"].apply(
                        lambda v: f"R$ {float(v):,.2f}" if str(v) not in ("","nan") else "")
                    df_ed_rede.insert(0, "✔", False)
                    edited_rede = st.data_editor(
                        df_ed_rede,
                        column_config={
                            "✔":            st.column_config.CheckboxColumn("✔",       width="small"),
                            "_idx":         None,   # oculta coluna técnica
                            "Status":       st.column_config.TextColumn("Status",      width="medium"),
                            "C.V. Rede":    st.column_config.TextColumn("C.V.",        width="medium"),
                            "Data Rede":    st.column_config.TextColumn("Data",        width="small"),
                            "Bandeira Rede":st.column_config.TextColumn("Bandeira",    width="small"),
                            "Valor Rede":   st.column_config.TextColumn("Valor",       width="medium"),
                        },
                        disabled=["_idx","Status","C.V. Rede","Data Rede","Bandeira Rede","Valor Rede"],
                        hide_index=True, use_container_width=True,
                        key="ed_vinc_rede",
                        height=min(400, 45 + len(df_ed_rede)*35),
                    )
                    # Lê _idx direto do editor — imune a dessincronias de índice
                    sel_idx_rede = edited_rede[edited_rede["✔"]]["_idx"].astype(int).tolist()
                    if sel_idx_rede:
                        total_sel_rede = df_conc_caixa[df_conc_caixa["_idx"].isin(sel_idx_rede)]["Valor Rede"].apply(
                            lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                        st.metric("Selecionado Rede", f"R$ {total_sel_rede:,.2f}",
                                  help=f"{len(sel_idx_rede)} registro(s)")

            # ── Comparação e confirmação ─────────
            st.markdown("---")
            obs_vinc = st.text_input("Observação (opcional):", key="obs_vinc_cx",
                                     placeholder="Ex: parcelado, data de liquidação diferente...")

            pode_vincular = len(sel_idx_cx) > 0 or len(sel_idx_rede) > 0
            if not pode_vincular:
                st.caption("Selecione ao menos um registro de cada lado para vincular.")
            else:
                # Mostra comparativo de valores
                val_cx   = df_conc_caixa[df_conc_caixa["_idx"].isin(sel_idx_cx)]["Valor Caixa"].apply(
                    lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                val_rede = df_conc_caixa[df_conc_caixa["_idx"].isin(sel_idx_rede)]["Valor Rede"].apply(
                    lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                diff = abs(val_cx - val_rede)
                diff_pct = (diff / val_cx * 100) if val_cx else 0

                mc1, mc2, mc3 = st.columns(3)
                mc1.metric("Total Caixa selecionado",  f"R$ {val_cx:,.2f}",
                           delta=f"{len(sel_idx_cx)} registro(s)")
                mc2.metric("Total Rede selecionado",   f"R$ {val_rede:,.2f}",
                           delta=f"{len(sel_idx_rede)} registro(s)")
                mc3.metric("Diferença",                f"R$ {diff:,.2f}",
                           delta=f"{diff_pct:.1f}%",
                           delta_color="off" if diff < 0.02 else "inverse")

                btn_label = "🔗 Confirmar Vínculo Manual"
                if st.button(btn_label, type="primary", key="btn_vinc_cx"):
                    st.session_state["vinculos_caixa"].append({
                        "idx_caixa": sel_idx_cx,
                        "idx_rede":  sel_idx_rede,
                        "obs":       obs_vinc,
                        "val_cx":    val_cx,
                        "val_rede":  val_rede,
                    })
                    st.success(f"🔗 Vínculo criado! ({len(sel_idx_cx)} Caixa + {len(sel_idx_rede)} Rede)")
                    st.rerun()

            # ── Lista de vínculos criados ────────
            if st.session_state["vinculos_caixa"]:
                st.divider()
                st.markdown(f"**Vínculos manuais criados: {len(st.session_state['vinculos_caixa'])}**")
                for i, v in enumerate(st.session_state["vinculos_caixa"]):
                    with st.expander(
                        f"🔗 Vínculo #{i+1} — Caixa R$ {v['val_cx']:,.2f} × Rede R$ {v['val_rede']:,.2f}"
                        + (f" | {v['obs']}" if v.get('obs') else ""), expanded=False
                    ):
                        vc1, vc2 = st.columns(2)
                        vc1.write(f"**Índices Caixa:** {v['idx_caixa']}")
                        vc2.write(f"**Índices Rede:** {v['idx_rede']}")
                        if st.button("🗑️ Remover vínculo", key=f"rm_vinc_{i}"):
                            st.session_state["vinculos_caixa"].pop(i)
                            st.rerun()

        # ── Export Excel da aba Caixa ──────────
        def exportar_caixa(df_caixa, df_conc, vinculos):
            out = io.BytesIO()
            df_conc_exp = df_conc.drop(columns=["_idx","forma_norm"], errors="ignore")
            with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
                df_caixa.drop(columns=["data_hora","forma_norm","bandeira_norm","data"], errors="ignore").to_excel(
                    writer, sheet_name="Caixa Completo", index=False)
                df_conc_exp.to_excel(writer, sheet_name="Conciliação Caixa", index=False)
                # Aba de vínculos manuais
                if vinculos:
                    rows_vinc = []
                    for i, v in enumerate(vinculos):
                        rows_vinc.append({
                            "Vínculo #":    i + 1,
                            "Índices Caixa": str(v["idx_caixa"]),
                            "Índices Rede":  str(v["idx_rede"]),
                            "Valor Caixa":   v["val_cx"],
                            "Valor Rede":    v["val_rede"],
                            "Diferença":     abs(v["val_cx"] - v["val_rede"]),
                            "Observação":    v.get("obs", ""),
                        })
                    pd.DataFrame(rows_vinc).to_excel(
                        writer, sheet_name="Vínculos Manuais Cx", index=False)
                wb     = writer.book
                fmt_h  = wb.add_format({"bold":True,"bg_color":"#1F3864","font_color":"white","border":1})
                fmt_ok = wb.add_format({"bg_color":"#C6EFCE"})
                fmt_w  = wb.add_format({"bg_color":"#FFEB9C"})
                fmt_man= wb.add_format({"bg_color":"#BDD7EE"})
                fmt_e  = wb.add_format({"bg_color":"#FFC7CE"})
                ws = writer.sheets["Conciliação Caixa"]
                for cn, col in enumerate(df_conc_exp.columns):
                    ws.write(0, cn, col, fmt_h); ws.set_column(cn, cn, 20)
                for rn in range(1, len(df_conc_exp)+1):
                    s = str(df_conc_exp.iloc[rn-1]["Status"])
                    ws.set_row(rn, None,
                               fmt_ok  if "✅" in s else
                               fmt_man if "🔗" in s else
                               fmt_w   if "⚠️" in s else fmt_e)
            return out.getvalue()

        excel_cx = exportar_caixa(df_caixa, df_conc_caixa, st.session_state["vinculos_caixa"])
        st.download_button(
            "⬇️ Exportar Caixa (Excel)",
            data=excel_cx,
            file_name=f"caixa_{filial_caixa}_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.divider()

# ════════════════════════════════════════════
# ABA 6 — PIX (POS + TEF)  — só renderiza se aba existe
# ════════════════════════════════════════════
with (aba_pix_pos if aba_pix_pos is not None else st.empty()):
    st.subheader("🔵 Conciliação PIX × OFX")

    tem_pos = bool(file_pix_pos)
    tem_tef = bool(file_pix_tef)

    if not tem_pos and not tem_tef:
        st.info("👈 Importe ao menos um arquivo PIX (POS ou TEF) na barra lateral.")
    else:
        df_pix_pos_data = pd.DataFrame()
        df_pix_tef_data = pd.DataFrame()
        meta_pos = {}
        meta_tef = {}

        # ── Parse POS ──
        if tem_pos:
            try:
                df_pix_pos_data, meta_pos = parse_pix_pos(file_pix_pos, df_estab)
            except Exception as e:
                st.error(f"Erro ao ler PIX POS: {e}")

        # ── Parse TEF ──
        if tem_tef:
            try:
                df_pix_tef_data, meta_tef = parse_pix_tef(file_pix_tef, df_estab)
            except Exception as e:
                st.error(f"Erro ao ler PIX TEF: {e}")

        # ── Metadata / info ──
        info_cols = st.columns(2)
        if tem_pos and meta_pos:
            acctid_pix = meta_pos.get("acctid","") or acctid_do_nome_arquivo_pix(getattr(file_pix_pos, "name", ""))
            with info_cols[0]:
                if meta_pos.get("formato") == "csv_sheets":
                    st.info(
                        f"**POS (Sheets)** — {meta_pos.get('total_filiais','?')} filiais  |  "
                        f"Período: {meta_pos.get('periodo_inicio','')} → {meta_pos.get('periodo_fim','')}  |  "
                        f"Total: {meta_pos.get('total_recebimentos','')}  |  "
                        f"Filiais: {meta_pos.get('filiais','')}"
                    )
                else:
                    st.info(f"**POS** — {meta_pos.get('nome','—')}  |  "
                            f"Ag: {meta_pos.get('agencia','—')} Cc: {meta_pos.get('conta','—')}  |  "
                            + (f"Período: {meta_pos.get('periodo','')}" if meta_pos.get('periodo') else ""))
        if tem_tef and meta_tef:
            with info_cols[1]:
                if meta_tef.get("formato") == "csv_sheets":
                    st.info(
                        f"**TEF (Sheets)** — {meta_tef.get('total_filiais','?')} filiais  |  "
                        f"Período: {meta_tef.get('periodo_inicio','')} → {meta_tef.get('periodo_fim','')}  |  "
                        f"Total: {meta_tef.get('total_recebimentos','')}  |  "
                        f"Filiais: {meta_tef.get('filiais','')}"
                    )
                else:
                    st.info(
                        f"**TEF** — {meta_tef.get('associado','—')}  |  "
                        f"Período: {meta_tef.get('periodo_inicio','')} → {meta_tef.get('periodo_fim','')}"
                    )

        # ── KPIs ──
        n_pos    = len(df_pix_pos_data)
        n_tef    = len(df_pix_tef_data)
        val_pos  = df_pix_pos_data["valor_pago_num"].sum() if n_pos else 0
        val_tef  = df_pix_tef_data["valor_num"].sum()      if n_tef else 0
        tar_pos  = df_pix_pos_data["tarifa_num"].sum()      if n_pos else 0

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("📟 POS — Transações",  n_pos)
        k2.metric("📱 TEF — Transações",  n_tef)
        k3.metric("💰 Valor Total POS",   f"R$ {val_pos:,.2f}")
        k4.metric("💰 Valor Total TEF",   f"R$ {val_tef:,.2f}")

        st.divider()

        # ── Conciliação unificada ──
        with st.spinner("Conciliando PIX com OFX..."):
            df_conc_pix, df_ofx_pix_usado = conciliar_pix_unificado(
                df_pix_pos_data, df_pix_tef_data, df_ofx_raw, tolerancia_dias=2
            )

        df_conc_pix = df_conc_pix.reset_index(drop=True)
        df_conc_pix["_idx"] = df_conc_pix.index

        # Aplica vínculos manuais (POS e TEF juntos)
        for vinc in st.session_state["vinculos_pix_pos"] + st.session_state["vinculos_pix_tef"]:
            for i in vinc.get("idx_pos", []) + vinc.get("idx_ofx", []):
                if i < len(df_conc_pix):
                    df_conc_pix.at[i, "Status"] = "🔗 Vinculado Manualmente"

        # Atualiza status PIX geral no session_state (para pivot do Caixa)
        _n_err = df_conc_pix["Status"].str.startswith("❌").sum()
        _n_man = df_conc_pix["Status"].str.startswith("🔗").sum()
        _n_ok  = df_conc_pix["Status"].str.startswith("✅").sum()
        _tot   = len(df_conc_pix)
        if _n_err > 0:    _spg = "❌"
        elif _n_man > 0:  _spg = "🔗"
        elif _n_ok  > 0:  _spg = "✅"
        else:             _spg = "—"
        st.session_state["_status_pix_geral"] = _spg

        # ── KPIs conciliação ──
        st.markdown("#### Resultado da Conciliação")
        p1, p2, p3, p4 = st.columns(4)
        p1.metric("📋 Total",           _tot)
        p2.metric("✅ Conciliados",      f"{_n_ok} ({_n_ok/_tot*100:.1f}%)" if _tot else "0")
        p3.metric("🔗 Vinc. Manual",     _n_man)
        p4.metric("❌ Não encontrados",  f"{_n_err} ({_n_err/_tot*100:.1f}%)" if _tot else "0")

        # ── Filtros ──
        fc1, fc2 = st.columns(2)
        with fc1:
            status_opts = sorted(df_conc_pix["Status"].unique().tolist())
            filtro_status = st.multiselect("Status:", status_opts, default=status_opts, key="filtro_pix_status")
        with fc2:
            origem_opts = sorted(df_conc_pix["Origem"].unique().tolist())
            filtro_origem = st.multiselect("Origem:", origem_opts, default=origem_opts, key="filtro_pix_origem")

        df_pix_show = df_conc_pix[
            df_conc_pix["Status"].isin(filtro_status) &
            df_conc_pix["Origem"].isin(filtro_origem)
        ].copy()

        # ── Tabela unificada ──
        df_pix_disp = df_pix_show.drop(columns=["_idx","_src_idx"], errors="ignore").copy()
        for c in ["Valor Pago","Tarifa","Valor OFX"]:
            df_pix_disp[c] = df_pix_disp[c].apply(
                lambda v: f"R$ {float(v):,.2f}" if str(v) not in ("","nan") else "")

        col_cfg_pix = {
            "Status":       st.column_config.TextColumn("Status",        width="medium"),
            "Origem":       st.column_config.TextColumn("Origem",        width="small"),
            "Filial":       st.column_config.TextColumn("Filial",        width="small"),
            "Data Pago":    st.column_config.TextColumn("Data Pago",     width="small"),
            "Identificador":st.column_config.TextColumn("ID/EndToEnd",   width="large"),
            "Pagador":      st.column_config.TextColumn("Pagador",       width="medium"),
            "Valor Pago":   st.column_config.TextColumn("Valor Pago",    width="medium"),
            "Tarifa":       st.column_config.TextColumn("Tarifa",        width="small"),
            "Data OFX":     st.column_config.TextColumn("Data OFX",      width="small"),
            "Memo OFX":     st.column_config.TextColumn("Memo OFX",      width="medium"),
            "Valor OFX":    st.column_config.TextColumn("Valor OFX",     width="medium"),
            "FITID OFX":    st.column_config.TextColumn("FITID OFX",     width="medium"),
        }
        st.dataframe(df_pix_disp, use_container_width=True, hide_index=True,
                     column_config=col_cfg_pix, height=420)

        st.divider()

        # ── Vinculação Manual ──
        st.markdown("#### 🔗 Vinculação Manual")

        mask_pend = (
            ~df_conc_pix["Status"].str.startswith("✅") &
            ~df_conc_pix["Status"].str.startswith("🔗")
        )
        df_pend = df_conc_pix[mask_pend].copy()

        if df_pend.empty:
            st.success("Todos os registros estão conciliados ou vinculados.")
        else:
            st.info(f"{len(df_pend)} registro(s) pendente(s).")

            df_pend_pix = df_pend[df_pend["Origem"].isin(["POS","TEF"])].copy()
            df_pend_ofx = df_pend[df_pend["Origem"] == "OFX"].copy()

            col_vl, col_vr = st.columns(2)

            with col_vl:
                st.markdown("**PIX POS/TEF sem par no OFX:**")
                if df_pend_pix.empty:
                    st.info("Nenhum pendente.")
                    sel_idx_pix = []
                else:
                    df_ed_pix = df_pend_pix[["_idx","Status","Origem","Data Pago","Pagador","Valor Pago"]].copy().reset_index(drop=True)
                    df_ed_pix["Valor Pago"] = df_ed_pix["Valor Pago"].apply(
                        lambda v: f"R$ {float(v):,.2f}" if str(v) not in ("","nan") else "")
                    df_ed_pix.insert(0, "✔", False)
                    edited_pix = st.data_editor(
                        df_ed_pix,
                        column_config={
                            "✔":        st.column_config.CheckboxColumn("✔",       width="small"),
                            "_idx":     None,
                            "Status":   st.column_config.TextColumn("Status",     width="medium"),
                            "Origem":   st.column_config.TextColumn("Origem",     width="small"),
                            "Data Pago":st.column_config.TextColumn("Data",       width="small"),
                            "Pagador":  st.column_config.TextColumn("Pagador",    width="medium"),
                            "Valor Pago":st.column_config.TextColumn("Valor",     width="medium"),
                        },
                        disabled=["_idx","Status","Origem","Data Pago","Pagador","Valor Pago"],
                        hide_index=True, use_container_width=True,
                        key="ed_vinc_pix_unified",
                        height=min(400, 45 + len(df_ed_pix)*35),
                    )
                    sel_idx_pix = edited_pix[edited_pix["✔"]]["_idx"].astype(int).tolist()
                    if sel_idx_pix:
                        tot_sel = df_conc_pix[df_conc_pix["_idx"].isin(sel_idx_pix)]["Valor Pago"].apply(
                            lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                        st.metric("Selecionado PIX", f"R$ {tot_sel:,.2f}",
                                  delta=f"{len(sel_idx_pix)} registro(s)")

            with col_vr:
                st.markdown("**OFX PIX sem par no POS/TEF:**")
                if df_pend_ofx.empty:
                    st.info("Nenhum pendente.")
                    sel_idx_ofx = []
                else:
                    df_ed_ofx = df_pend_ofx[["_idx","Status","Data OFX","Memo OFX","Valor OFX"]].copy().reset_index(drop=True)
                    df_ed_ofx["Valor OFX"] = df_ed_ofx["Valor OFX"].apply(
                        lambda v: f"R$ {float(v):,.2f}" if str(v) not in ("","nan") else "")
                    df_ed_ofx.insert(0, "✔", False)
                    edited_ofx = st.data_editor(
                        df_ed_ofx,
                        column_config={
                            "✔":        st.column_config.CheckboxColumn("✔",     width="small"),
                            "_idx":     None,
                            "Status":   st.column_config.TextColumn("Status",   width="medium"),
                            "Data OFX": st.column_config.TextColumn("Data",     width="small"),
                            "Memo OFX": st.column_config.TextColumn("Memo",     width="large"),
                            "Valor OFX":st.column_config.TextColumn("Valor",    width="medium"),
                        },
                        disabled=["_idx","Status","Data OFX","Memo OFX","Valor OFX"],
                        hide_index=True, use_container_width=True,
                        key="ed_vinc_ofx_pix_unified",
                        height=min(400, 45 + len(df_ed_ofx)*35),
                    )
                    sel_idx_ofx = edited_ofx[edited_ofx["✔"]]["_idx"].astype(int).tolist()
                    if sel_idx_ofx:
                        tot_ofx = df_conc_pix[df_conc_pix["_idx"].isin(sel_idx_ofx)]["Valor OFX"].apply(
                            lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                        st.metric("Selecionado OFX", f"R$ {tot_ofx:,.2f}",
                                  delta=f"{len(sel_idx_ofx)} registro(s)")

            st.markdown("---")
            obs_pix_vinc = st.text_input("Observação (opcional):", key="obs_vinc_pix_u",
                                          placeholder="Ex: PIX recebido fora do período...")

            pode_vincular = len(sel_idx_pix) > 0 or len(sel_idx_ofx) > 0
            if pode_vincular:
                val_pix_sel = df_conc_pix[df_conc_pix["_idx"].isin(sel_idx_pix)]["Valor Pago"].apply(
                    lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                val_ofx_sel = df_conc_pix[df_conc_pix["_idx"].isin(sel_idx_ofx)]["Valor OFX"].apply(
                    lambda v: float(v) if str(v) not in ("","nan") else 0).sum()
                diff_vinc = abs(val_pix_sel - val_ofx_sel)

                mc1, mc2, mc3 = st.columns(3)
                mc1.metric("Total PIX selecionado", f"R$ {val_pix_sel:,.2f}",
                           delta=f"{len(sel_idx_pix)} registro(s)")
                mc2.metric("Total OFX selecionado", f"R$ {val_ofx_sel:,.2f}",
                           delta=f"{len(sel_idx_ofx)} registro(s)")
                mc3.metric("Diferença", f"R$ {diff_vinc:,.2f}",
                           delta_color="off" if diff_vinc < 0.02 else "inverse")

                if st.button("🔗 Confirmar Vínculo Manual PIX", type="primary", key="btn_vinc_pix_u"):
                    st.session_state["vinculos_pix_pos"].append({
                        "idx_pos": sel_idx_pix,
                        "idx_ofx": sel_idx_ofx,
                        "obs":     obs_pix_vinc,
                        "val_pos": val_pix_sel,
                        "val_ofx": val_ofx_sel,
                    })
                    st.success("🔗 Vínculo criado!")
                    st.rerun()
            else:
                st.caption("Selecione ao menos um registro para vincular.")

            # Lista de vínculos
            todos_vinc = (
                [(v, "POS/TEF") for v in st.session_state["vinculos_pix_pos"]] +
                [(v, "TEF")     for v in st.session_state["vinculos_pix_tef"]]
            )
            if todos_vinc:
                st.divider()
                st.markdown(f"**Vínculos manuais: {len(todos_vinc)}**")
                for i, (v, origem) in enumerate(todos_vinc):
                    with st.expander(
                        f"🔗 Vínculo #{i+1} ({origem}) — "
                        f"PIX R$ {v.get('val_pos',0):,.2f} × OFX R$ {v.get('val_ofx',0):,.2f}"
                        + (f" | {v['obs']}" if v.get('obs') else ""), expanded=False
                    ):
                        vc1, vc2 = st.columns(2)
                        vc1.write(f"**Índices PIX:** {v.get('idx_pos',[])}")
                        vc2.write(f"**Índices OFX:** {v.get('idx_ofx',[])}")
                        if st.button("🗑️ Remover", key=f"rm_pix_u_{i}"):
                            if origem == "TEF":
                                st.session_state["vinculos_pix_tef"].pop(i - len(st.session_state["vinculos_pix_pos"]))
                            else:
                                st.session_state["vinculos_pix_pos"].pop(i)
                            st.rerun()

        # ── Export ──
        def exportar_pix_unificado(df_conc, meta_p, meta_t, vinc_pos, vinc_tef):
            out = io.BytesIO()
            df_exp = df_conc.drop(columns=["_idx","_src_idx"], errors="ignore")
            with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
                df_exp.to_excel(writer, sheet_name="Conciliação PIX", index=False)
                todos = vinc_pos + vinc_tef
                if todos:
                    pd.DataFrame([{
                        "Vínculo #":    i+1,
                        "Índices PIX":  str(v.get("idx_pos",[])),
                        "Índices OFX":  str(v.get("idx_ofx",[])),
                        "Valor PIX":    v.get("val_pos",0),
                        "Valor OFX":    v.get("val_ofx",0),
                        "Diferença":    abs(v.get("val_pos",0)-v.get("val_ofx",0)),
                        "Observação":   v.get("obs",""),
                    } for i, v in enumerate(todos)]).to_excel(
                        writer, sheet_name="Vínculos Manuais PIX", index=False)
                wb   = writer.book
                fmt_h   = wb.add_format({"bold":True,"bg_color":"#1F3864","font_color":"white","border":1})
                fmt_ok  = wb.add_format({"bg_color":"#C6EFCE"})
                fmt_man = wb.add_format({"bg_color":"#BDD7EE"})
                fmt_e   = wb.add_format({"bg_color":"#FFC7CE"})
                ws2 = writer.sheets["Conciliação PIX"]
                for cn, col in enumerate(df_exp.columns):
                    ws2.write(0, cn, col, fmt_h); ws2.set_column(cn, cn, 22)
                for rn in range(1, len(df_exp)+1):
                    s = str(df_exp.iloc[rn-1]["Status"])
                    ws2.set_row(rn, None, fmt_ok if "✅" in s else (fmt_man if "🔗" in s else fmt_e))
            return out.getvalue()

        excel_pix = exportar_pix_unificado(
            df_conc_pix, meta_pos, meta_tef,
            st.session_state["vinculos_pix_pos"],
            st.session_state["vinculos_pix_tef"],
        )
        st.download_button(
            "⬇️ Exportar PIX (Excel)",
            data=excel_pix,
            file_name=f"pix_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


st.divider()

# ─────────────────────────────────────────────
# EXPORTAR
# ─────────────────────────────────────────────
df_detalhe_export = build_status_transacao(
    df_rede_orig, df_rede_grupo, df_result,
    st.session_state["vinculos_manuais"]
)
excel_bytes = exportar_excel(
    df_result, df_detalhe_export, df_rede_grupo,
    st.session_state["vinculos_manuais"]
)
st.download_button(
    label="⬇️ Exportar Resultado Completo (Excel)",
    data=excel_bytes,
    file_name=f"conciliacao_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
